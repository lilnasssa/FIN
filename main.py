# =============================================================================
#  22-17 LABEL FINANCE BOT — Bothost edition
#  Управление финансами лейбла: артисты, доходы/расходы, Excel-импорт,
#  отчёты за любой период, выплаты и поиск аномалий.
#  ИИ: Qwen 3.8 Max по умолчанию + автоэскалация к Sonnet 5 / GPT 5.5 /
#  Opus 5 / GPT 5.6 Luna / Terra / Sol через Xkiro и Tooken.
#
#  Интерфейс: ТОЛЬКО inline-кнопки, без slash-команд.
#
#  Сборка под Bothost:
#   * точка входа — main.py, зависимости — requirements.txt;
#   * БД по умолчанию SQLite в /app/data (единственная папка, которая
#     переживает передеплой). Если задан DATABASE_URL с postgres:// —
#     автоматически используется PostgreSQL (psycopg2), код общий;
#   * все секреты — из переменных окружения панели Bothost;
#   * long polling, один процесс, без вебхуков и открытых портов.
# =============================================================================
from __future__ import annotations

import asyncio
import csv
import hashlib
import json
import logging
import os
import re
import secrets
import sqlite3
import statistics
import threading
import time
from contextlib import contextmanager
from contextvars import ContextVar
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Iterator, Sequence

from telegram import (
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    InputFile,
    Update,
)
from telegram.constants import ChatAction, ParseMode
from telegram.error import BadRequest
from telegram.ext import (
    ApplicationBuilder,
    CallbackQueryHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

try:
    from telegram.ext import AIORateLimiter
except ImportError:
    AIORateLimiter = None

# -----------------------------------------------------------------------------
# Конфигурация и логирование
# -----------------------------------------------------------------------------
logging.basicConfig(
    format="%(asctime)s | %(levelname)-8s | %(name)s | %(message)s",
    level=os.getenv("LOG_LEVEL", "INFO"),
)
logging.getLogger("httpx").setLevel(logging.WARNING)
log = logging.getLogger("finbot")


def _env_decimal(name: str, default: str) -> Decimal:
    try:
        return Decimal(os.getenv(name, default).replace(",", "."))
    except InvalidOperation:
        return Decimal(default)


# Метка сборки: видна в логах и в главном меню. Нужна, чтобы отличать,
# какая версия реально задеплоена на хостинге.
BUILD = "2026-08-18 multi-model"
# Имя модели, которая реально отвечает сейчас (заполняется при автоподборе).
ACTIVE_MODEL: dict[str, str] = {}


class Config:
    BOT_TOKEN: str = os.getenv("BOT_TOKEN", "").strip()
    # ---------- ИИ: провайдеры (OpenAI-совместимый API) ----------
    # Xkiro
    XKIRO_API_KEY: str = os.getenv("XKIRO_API_KEY", "").strip()
    XKIRO_BASE_URL: str = (
        os.getenv("XKIRO_BASE_URL", "https://api.xkiro.ru/v1").strip().rstrip("/")
    )
    # Tooken (https://tooken.club)
    TOOKEN_API_KEY: str = os.getenv("TOOKEN_API_KEY", "").strip()
    TOOKEN_BASE_URL: str = (
        os.getenv("TOOKEN_BASE_URL", "https://tooken.club/v1").strip().rstrip("/")
    )
    # Порядок обхода: первый рабочий отвечает, остальные — резерв.
    AI_PROVIDER_ORDER: str = os.getenv("AI_PROVIDER_ORDER", "xkiro,tooken").strip()

    # ---------- ИИ: модели и маршрутизация ----------
    AI_MODE: str = os.getenv("AI_MODE", "auto").strip().lower()  # auto | manual
    AI_BASE_MODEL: str = os.getenv("AI_BASE_MODEL", "qwen").strip().lower()
    AI_MANUAL_MODEL: str = os.getenv("AI_MANUAL_MODEL", "qwen").strip().lower()
    # Разрешать ли самые дорогие модели (Terra/Sol) в крайних случаях.
    AI_ALLOW_EXTREME: bool = os.getenv("AI_ALLOW_EXTREME", "1").strip().lower() not in (
        "0",
        "false",
        "no",
    )
    # Ниже этого уровня уверенности задача уходит следующей модели.
    AI_CONFIDENCE_MIN: Decimal = _env_decimal("AI_CONFIDENCE_MIN", "0.72")
    AI_MAX_TOKENS: int = int(os.getenv("AI_MAX_TOKENS", "1100"))
    # Для больших отчётов — отдельный (чуть больше) потолок.
    AI_MAX_TOKENS_LONG: int = int(os.getenv("AI_MAX_TOKENS_LONG", "1600"))
    # Режим экономии: короткие промпты, мало шагов, кэш ответов.
    AI_ECONOMY: bool = os.getenv("AI_ECONOMY", "1").strip().lower() not in (
        "0",
        "false",
        "no",
    )
    # Сколько моделей максимум может участвовать в одном запросе.
    AI_MAX_STEPS: int = int(os.getenv("AI_MAX_STEPS", "2"))
    # Повторные одинаковые запросы берём из кэша (секунды).
    AI_CACHE_TTL: int = int(os.getenv("AI_CACHE_TTL", "21600"))
    AI_CACHE_SIZE: int = int(os.getenv("AI_CACHE_SIZE", "64"))
    # Суточный лимит токенов: после него только дешёвые модели (0 = без лимита).
    AI_DAILY_TOKEN_LIMIT: int = int(os.getenv("AI_DAILY_TOKEN_LIMIT", "300000"))
    # Сколько строк данных максимум уезжает в модель.
    AI_MAX_DATA_ROWS: int = int(os.getenv("AI_MAX_DATA_ROWS", "25"))
    AI_MAX_EXCEL_AI_ROWS: int = int(os.getenv("AI_MAX_EXCEL_AI_ROWS", "120"))
    AI_TIMEOUT: int = int(os.getenv("AI_TIMEOUT", "120"))
    AI_TEMPERATURE: Decimal = _env_decimal("AI_TEMPERATURE", "0.2")
    # Реальные id моделей у провайдера — меняются без правки кода.
    MODEL_QWEN: str = os.getenv("MODEL_QWEN", "qwen-3.8-max").strip()
    MODEL_SONNET5: str = os.getenv("MODEL_SONNET5", "claude-sonnet-5").strip()
    MODEL_OPUS5: str = os.getenv("MODEL_OPUS5", "claude-opus-5").strip()
    MODEL_GPT55: str = os.getenv("MODEL_GPT55", "gpt-5.5").strip()
    MODEL_LUNA: str = os.getenv("MODEL_LUNA", "gpt-5.6-luna").strip()
    MODEL_TERRA: str = os.getenv("MODEL_TERRA", "gpt-5.6-terra").strip()
    MODEL_SOL: str = os.getenv("MODEL_SOL", "gpt-5.6-sol").strip()

    # Пусто => SQLite в DATA_DIR. postgres://... => PostgreSQL.
    DATABASE_URL: str = os.getenv("DATABASE_URL", "").strip()
    DATA_DIR: str = os.getenv("DATA_DIR", "/app/data")
    SQLITE_NAME: str = os.getenv("SQLITE_NAME", "finbot.db")
    DB_POOL_MIN: int = int(os.getenv("DB_POOL_MIN", "1"))
    DB_POOL_MAX: int = int(os.getenv("DB_POOL_MAX", "5"))

    ALLOWED_USER_IDS: set[int] = {
        int(x)
        for x in os.getenv("ALLOWED_USER_IDS", "").replace(" ", "").split(",")
        if x.strip().lstrip("-").isdigit()
    }
    # Владельцы: полные права, включая управление доступом. Задаются один раз.
    OWNER_IDS: set[int] = {
        int(x)
        for x in os.getenv("OWNER_IDS", "").replace(" ", "").split(",")
        if x.strip().lstrip("-").isdigit()
    }
    # Срок жизни кода-приглашения в часах (0 = бессрочно) и число активаций.
    INVITE_TTL_HOURS: int = int(os.getenv("INVITE_TTL_HOURS", "72"))
    INVITE_DEFAULT_USES: int = int(os.getenv("INVITE_DEFAULT_USES", "1"))

    DEFAULT_CURRENCY: str = os.getenv("DEFAULT_CURRENCY", "RUB")
    DEFAULT_RATE: Decimal = _env_decimal("DEFAULT_RATE", "0.20")
    MAX_EXCEL_BYTES: int = int(os.getenv("MAX_EXCEL_MB", "10")) * 1024 * 1024
    MAX_EXCEL_ROWS: int = int(os.getenv("MAX_EXCEL_ROWS", "500"))
    ANOMALY_SIGMA: Decimal = _env_decimal("ANOMALY_SIGMA", "2")
    ANOMALY_MONTHS: int = int(os.getenv("ANOMALY_MONTHS", "6"))
    LABEL_NAME: str = os.getenv("LABEL_NAME", "22-17")
    # Ставка налога для резерва в сводке (0.06 = УСН 6% с дохода).
    TAX_RATE: Decimal = _env_decimal("TAX_RATE", "0.06")
    # Автоматическая сводка 1-го числа владельцам и админам.
    MONTHLY_DIGEST: bool = os.getenv("MONTHLY_DIGEST", "1").strip() not in (
        "0",
        "false",
        "False",
        "",
    )
    DIGEST_HOUR: int = int(os.getenv("DIGEST_HOUR", "10"))


# -----------------------------------------------------------------------------
# Деньги: храним в копейках (целые числа), считаем в Decimal
# -----------------------------------------------------------------------------
CENTS = Decimal("0.01")


def to_cents(value: Any) -> int:
    return int(
        (Decimal(str(value)).quantize(CENTS, rounding=ROUND_HALF_UP) * 100)
        .to_integral_value(rounding=ROUND_HALF_UP)
    )


def from_cents(value: Any) -> Decimal:
    return (Decimal(int(value or 0)) / 100).quantize(CENTS)


def as_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()


def json_dumps(obj: Any) -> str:
    def default(o: Any) -> Any:
        if isinstance(o, Decimal):
            return float(o)
        if isinstance(o, (date, datetime)):
            return o.isoformat()
        return str(o)

    return json.dumps(obj, ensure_ascii=False, default=default)


# -----------------------------------------------------------------------------
# Схемы БД
# -----------------------------------------------------------------------------
SCHEMA_SQLITE = """
CREATE TABLE IF NOT EXISTS artists (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    name            TEXT NOT NULL UNIQUE,
    rate            REAL NOT NULL DEFAULT 0.2,
    tg_username     TEXT,
    is_active       INTEGER NOT NULL DEFAULT 1,
    created_at      TEXT NOT NULL DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS transactions (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    artist_id       INTEGER REFERENCES artists(id) ON DELETE SET NULL,
    kind            TEXT NOT NULL CHECK (kind IN ('income', 'expense')),
    amount_cents    INTEGER NOT NULL CHECK (amount_cents > 0),
    currency        TEXT NOT NULL DEFAULT 'RUB',
    category        TEXT,
    description     TEXT,
    occurred_on     TEXT NOT NULL,
    source          TEXT NOT NULL DEFAULT 'manual',
    external_key    TEXT,
    created_by      INTEGER,
    created_at      TEXT NOT NULL DEFAULT (datetime('now'))
);

CREATE UNIQUE INDEX IF NOT EXISTS transactions_external_key_uniq
    ON transactions (external_key) WHERE external_key IS NOT NULL;
CREATE INDEX IF NOT EXISTS transactions_occurred_idx ON transactions (occurred_on);
CREATE INDEX IF NOT EXISTS transactions_artist_idx ON transactions (artist_id);

CREATE TABLE IF NOT EXISTS payments (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    artist_id       INTEGER NOT NULL REFERENCES artists(id) ON DELETE CASCADE,
    period_start    TEXT NOT NULL,
    period_end      TEXT NOT NULL,
    revenue_cents   INTEGER NOT NULL DEFAULT 0,
    expenses_cents  INTEGER NOT NULL DEFAULT 0,
    rate            REAL NOT NULL,
    amount_cents    INTEGER NOT NULL,
    status          TEXT NOT NULL DEFAULT 'pending'
                    CHECK (status IN ('pending', 'paid', 'canceled')),
    paid_at         TEXT,
    created_at      TEXT NOT NULL DEFAULT (datetime('now')),
    UNIQUE (artist_id, period_start, period_end)
);

CREATE TABLE IF NOT EXISTS ai_reports (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    kind            TEXT NOT NULL,
    period_start    TEXT,
    period_end      TEXT,
    payload         TEXT,
    body            TEXT,
    created_at      TEXT NOT NULL DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS audit_log (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    tg_user_id      INTEGER,
    action          TEXT NOT NULL,
    details         TEXT,
    created_at      TEXT NOT NULL DEFAULT (datetime('now'))
);

CREATE TABLE IF NOT EXISTS bot_users (
    tg_user_id      INTEGER PRIMARY KEY,
    username        TEXT,
    full_name       TEXT,
    role            TEXT NOT NULL DEFAULT 'viewer'
                    CHECK (role IN ('owner', 'admin', 'viewer')),
    is_active       INTEGER NOT NULL DEFAULT 1,
    invited_by      INTEGER,
    created_at      TEXT NOT NULL DEFAULT (datetime('now')),
    last_seen_at    TEXT
);

CREATE TABLE IF NOT EXISTS invites (
    code            TEXT PRIMARY KEY,
    role            TEXT NOT NULL DEFAULT 'viewer'
                    CHECK (role IN ('admin', 'viewer')),
    uses_left       INTEGER NOT NULL DEFAULT 1,
    expires_at      TEXT,
    created_by      INTEGER,
    is_revoked      INTEGER NOT NULL DEFAULT 0,
    created_at      TEXT NOT NULL DEFAULT (datetime('now'))
);
"""

SCHEMA_POSTGRES = """
CREATE TABLE IF NOT EXISTS artists (
    id              SERIAL PRIMARY KEY,
    name            TEXT NOT NULL UNIQUE,
    rate            DOUBLE PRECISION NOT NULL DEFAULT 0.2,
    tg_username     TEXT,
    is_active       SMALLINT NOT NULL DEFAULT 1,
    created_at      TIMESTAMPTZ NOT NULL DEFAULT now()
);

CREATE TABLE IF NOT EXISTS transactions (
    id              BIGSERIAL PRIMARY KEY,
    artist_id       INTEGER REFERENCES artists(id) ON DELETE SET NULL,
    kind            TEXT NOT NULL CHECK (kind IN ('income', 'expense')),
    amount_cents    BIGINT NOT NULL CHECK (amount_cents > 0),
    currency        TEXT NOT NULL DEFAULT 'RUB',
    category        TEXT,
    description     TEXT,
    occurred_on     DATE NOT NULL,
    source          TEXT NOT NULL DEFAULT 'manual',
    external_key    TEXT,
    created_by      BIGINT,
    created_at      TIMESTAMPTZ NOT NULL DEFAULT now()
);

CREATE UNIQUE INDEX IF NOT EXISTS transactions_external_key_uniq
    ON transactions (external_key) WHERE external_key IS NOT NULL;
CREATE INDEX IF NOT EXISTS transactions_occurred_idx ON transactions (occurred_on);
CREATE INDEX IF NOT EXISTS transactions_artist_idx ON transactions (artist_id);

CREATE TABLE IF NOT EXISTS payments (
    id              BIGSERIAL PRIMARY KEY,
    artist_id       INTEGER NOT NULL REFERENCES artists(id) ON DELETE CASCADE,
    period_start    DATE NOT NULL,
    period_end      DATE NOT NULL,
    revenue_cents   BIGINT NOT NULL DEFAULT 0,
    expenses_cents  BIGINT NOT NULL DEFAULT 0,
    rate            DOUBLE PRECISION NOT NULL,
    amount_cents    BIGINT NOT NULL,
    status          TEXT NOT NULL DEFAULT 'pending'
                    CHECK (status IN ('pending', 'paid', 'canceled')),
    paid_at         TIMESTAMPTZ,
    created_at      TIMESTAMPTZ NOT NULL DEFAULT now(),
    UNIQUE (artist_id, period_start, period_end)
);

CREATE TABLE IF NOT EXISTS ai_reports (
    id              BIGSERIAL PRIMARY KEY,
    kind            TEXT NOT NULL,
    period_start    DATE,
    period_end      DATE,
    payload         TEXT,
    body            TEXT,
    created_at      TIMESTAMPTZ NOT NULL DEFAULT now()
);

CREATE TABLE IF NOT EXISTS audit_log (
    id              BIGSERIAL PRIMARY KEY,
    tg_user_id      BIGINT,
    action          TEXT NOT NULL,
    details         TEXT,
    created_at      TIMESTAMPTZ NOT NULL DEFAULT now()
);

CREATE TABLE IF NOT EXISTS bot_users (
    tg_user_id      BIGINT PRIMARY KEY,
    username        TEXT,
    full_name       TEXT,
    role            TEXT NOT NULL DEFAULT 'viewer'
                    CHECK (role IN ('owner', 'admin', 'viewer')),
    is_active       SMALLINT NOT NULL DEFAULT 1,
    invited_by      BIGINT,
    created_at      TIMESTAMPTZ NOT NULL DEFAULT now(),
    last_seen_at    TEXT
);

CREATE TABLE IF NOT EXISTS invites (
    code            TEXT PRIMARY KEY,
    role            TEXT NOT NULL DEFAULT 'viewer'
                    CHECK (role IN ('admin', 'viewer')),
    uses_left       INTEGER NOT NULL DEFAULT 1,
    expires_at      TEXT,
    created_by      BIGINT,
    is_revoked      SMALLINT NOT NULL DEFAULT 0,
    created_at      TIMESTAMPTZ NOT NULL DEFAULT now()
);
"""


class Database:
    """Единый слой доступа: SQLite (по умолчанию) или PostgreSQL.

    SQL пишется с плейсхолдерами `?`; для PostgreSQL они заменяются на `%s`.
    Все вызовы синхронные и идут из хендлеров через asyncio.to_thread.
    """

    def __init__(self) -> None:
        url = Config.DATABASE_URL
        self.is_postgres = url.startswith(("postgres://", "postgresql://"))

        if self.is_postgres:
            import psycopg2
            import psycopg2.extras
            from psycopg2.pool import ThreadedConnectionPool

            self._extras = psycopg2.extras
            self._pool = ThreadedConnectionPool(
                Config.DB_POOL_MIN, Config.DB_POOL_MAX, dsn=url
            )
            log.info("DB backend: PostgreSQL")
        else:
            directory = Path(Config.DATA_DIR)
            try:
                directory.mkdir(parents=True, exist_ok=True)
                probe = directory / ".write_test"
                probe.write_text("ok", encoding="utf-8")
                probe.unlink()
            except OSError:
                directory = Path("./data")
                directory.mkdir(parents=True, exist_ok=True)
                log.warning(
                    "%s недоступна для записи, использую %s "
                    "(данные не переживут передеплой!)",
                    Config.DATA_DIR,
                    directory.resolve(),
                )
            self.sqlite_path = str(directory / Config.SQLITE_NAME)
            self._lock = threading.RLock()
            self._conn = sqlite3.connect(
                self.sqlite_path, check_same_thread=False, timeout=30
            )
            self._conn.row_factory = sqlite3.Row
            self._conn.execute("PRAGMA journal_mode=WAL;")
            self._conn.execute("PRAGMA foreign_keys=ON;")
            log.info("DB backend: SQLite (%s)", self.sqlite_path)

    # ---------- утилиты диалекта ----------
    def _sql(self, sql: str) -> str:
        return sql.replace("?", "%s") if self.is_postgres else sql

    def d(self, value: date) -> Any:
        return value if self.is_postgres else value.isoformat()

    def now(self) -> Any:
        return (
            datetime.now()
            if self.is_postgres
            else datetime.now().isoformat(" ", "seconds")
        )

    @contextmanager
    def _pg_cursor(self, commit: bool) -> Iterator[Any]:
        conn = self._pool.getconn()
        try:
            with conn:
                with conn.cursor(cursor_factory=self._extras.RealDictCursor) as cur:
                    yield cur
                if not commit:
                    conn.rollback()
        except Exception:
            conn.rollback()
            raise
        finally:
            self._pool.putconn(conn)

    # ---------- операции ----------
    def query(self, sql: str, params: Sequence[Any] = ()) -> list[dict]:
        if self.is_postgres:
            with self._pg_cursor(commit=False) as cur:
                cur.execute(self._sql(sql), tuple(params))
                return [dict(r) for r in cur.fetchall()]
        with self._lock:
            cur = self._conn.execute(sql, tuple(params))
            return [dict(r) for r in cur.fetchall()]

    def scalar(self, sql: str, params: Sequence[Any] = ()) -> Any:
        rows = self.query(sql, params)
        if not rows:
            return None
        return next(iter(rows[0].values()))

    def execute(self, sql: str, params: Sequence[Any] = ()) -> int:
        if self.is_postgres:
            with self._pg_cursor(commit=True) as cur:
                cur.execute(self._sql(sql), tuple(params))
                return cur.rowcount
        with self._lock:
            cur = self._conn.execute(sql, tuple(params))
            self._conn.commit()
            return cur.rowcount

    def executemany(self, sql: str, rows: Sequence[Sequence[Any]]) -> int:
        if not rows:
            return 0
        if self.is_postgres:
            with self._pg_cursor(commit=True) as cur:
                self._extras.execute_batch(
                    cur, self._sql(sql), [tuple(r) for r in rows], page_size=200
                )
                return len(rows)
        with self._lock:
            self._conn.executemany(sql, [tuple(r) for r in rows])
            self._conn.commit()
            return len(rows)

    def init_schema(self) -> None:
        schema = SCHEMA_POSTGRES if self.is_postgres else SCHEMA_SQLITE
        if self.is_postgres:
            with self._pg_cursor(commit=True) as cur:
                cur.execute(schema)
        else:
            with self._lock:
                self._conn.executescript(schema)
                self._conn.commit()
        log.info("DB schema ready")

    def close(self) -> None:
        if self.is_postgres:
            self._pool.closeall()
        else:
            with self._lock:
                self._conn.close()


# -----------------------------------------------------------------------------
# Репозиторий
# -----------------------------------------------------------------------------
NO_ARTIST = "— без артиста —"
NO_CATEGORY = "без категории"


class Repo:
    def __init__(self, db: Database) -> None:
        self.db = db

    # ---------- артисты ----------
    def add_artist(self, name: str, rate: Decimal, username: str | None = None) -> dict:
        self.db.execute(
            """
            INSERT INTO artists (name, rate, tg_username, is_active)
            VALUES (?, ?, ?, 1)
            ON CONFLICT (name) DO UPDATE SET
                rate = excluded.rate,
                tg_username = COALESCE(excluded.tg_username, artists.tg_username),
                is_active = 1
            """,
            (name, float(rate), username),
        )
        rows = self.db.query(
            "SELECT id, name, rate, tg_username FROM artists WHERE name = ?", (name,)
        )
        return rows[0]

    def list_artists(self, only_active: bool = True) -> list[dict]:
        sql = "SELECT id, name, rate, tg_username, is_active FROM artists"
        if only_active:
            sql += " WHERE is_active = 1"
        return self.db.query(sql + " ORDER BY name")

    def get_artist(self, artist_id: int) -> dict | None:
        rows = self.db.query("SELECT * FROM artists WHERE id = ?", (artist_id,))
        return rows[0] if rows else None

    def resolve_artist_ids(self, names: list[str]) -> dict[str, int]:
        if not names:
            return {}
        placeholders = ", ".join(["?"] * len(names))
        rows = self.db.query(
            f"SELECT id, name FROM artists WHERE lower(name) IN ({placeholders})",
            [n.lower() for n in names],
        )
        return {str(r["name"]).lower(): int(r["id"]) for r in rows}

    # ---------- транзакции ----------
    INSERT_TX = """
        INSERT INTO transactions
            (artist_id, kind, amount_cents, currency, category, description,
             occurred_on, source, external_key, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT DO NOTHING
    """

    def add_transaction(
        self,
        artist_id: int | None,
        kind: str,
        amount: Decimal,
        currency: str,
        category: str | None,
        description: str | None,
        occurred_on: date,
        source: str = "manual",
        external_key: str | None = None,
        created_by: int | None = None,
    ) -> int:
        return self.db.execute(
            self.INSERT_TX,
            (
                artist_id,
                kind,
                to_cents(amount),
                currency,
                category,
                description,
                self.db.d(occurred_on),
                source,
                external_key,
                created_by,
            ),
        )

    def bulk_insert_transactions(self, rows: list[dict], created_by: int | None) -> int:
        if not rows:
            return 0
        before = int(self.db.scalar("SELECT COUNT(*) AS c FROM transactions") or 0)
        payload = [
            (
                r.get("artist_id"),
                r["kind"],
                to_cents(r["amount"]),
                (r.get("currency") or Config.DEFAULT_CURRENCY),
                r.get("category"),
                r.get("description"),
                self.db.d(r["occurred_on"]),
                "excel",
                r.get("external_key"),
                created_by,
            )
            for r in rows
        ]
        self.db.executemany(self.INSERT_TX, payload)
        after = int(self.db.scalar("SELECT COUNT(*) AS c FROM transactions") or 0)
        return max(after - before, 0)

    # ---------- аналитика ----------
    @staticmethod
    def month_bounds(year: int, month: int) -> tuple[date, date]:
        start = date(year, month, 1)
        end = date(year + (month == 12), (month % 12) + 1, 1)
        return start, end

    def monthly_breakdown(self, year: int, month: int) -> list[dict]:
        start, end = self.month_bounds(year, month)
        return self.breakdown(start, end)

    def breakdown(self, start: date, end: date) -> list[dict]:
        """Разбор по артистам за любой период (end не включается)."""
        rows = self.db.query(
            f"""
            SELECT COALESCE(a.name, '{NO_ARTIST}') AS artist,
                   COALESCE(a.rate, 0) AS rate,
                   SUM(CASE WHEN t.kind = 'income'  THEN t.amount_cents ELSE 0 END) AS revenue_cents,
                   SUM(CASE WHEN t.kind = 'expense' THEN t.amount_cents ELSE 0 END) AS expenses_cents,
                   COUNT(*) AS tx_count
            FROM transactions t
            LEFT JOIN artists a ON a.id = t.artist_id
            WHERE t.occurred_on >= ? AND t.occurred_on < ?
            GROUP BY COALESCE(a.name, '{NO_ARTIST}'), COALESCE(a.rate, 0)
            ORDER BY 3 DESC
            """,
            (self.db.d(start), self.db.d(end)),
        )
        result: list[dict] = []
        for r in rows:
            revenue = from_cents(r["revenue_cents"])
            expenses = from_cents(r["expenses_cents"])
            rate = Decimal(str(r["rate"] or 0))
            gross = revenue - expenses
            # Прибыль артиста = (Выручка - Расходы) * Ставка
            share = (gross * rate).quantize(CENTS, rounding=ROUND_HALF_UP)
            result.append(
                {
                    "artist": r["artist"],
                    "rate": rate,
                    "revenue": revenue,
                    "expenses": expenses,
                    "gross": gross,
                    "artist_share": share,
                    "label_profit": (gross - share).quantize(CENTS),
                    "tx_count": int(r["tx_count"]),
                }
            )
        return result

    def category_breakdown(self, year: int, month: int) -> list[dict]:
        start, end = self.month_bounds(year, month)
        return self.categories(start, end)

    def categories(self, start: date, end: date) -> list[dict]:
        """Статьи доходов/расходов за любой период."""
        rows = self.db.query(
            f"""
            SELECT kind,
                   COALESCE(category, '{NO_CATEGORY}') AS category,
                   SUM(amount_cents) AS total_cents,
                   COUNT(*) AS cnt
            FROM transactions
            WHERE occurred_on >= ? AND occurred_on < ?
            GROUP BY kind, COALESCE(category, '{NO_CATEGORY}')
            ORDER BY 3 DESC
            """,
            (self.db.d(start), self.db.d(end)),
        )
        return [
            {
                "kind": r["kind"],
                "category": r["category"],
                "total": from_cents(r["total_cents"]),
                "count": int(r["cnt"]),
            }
            for r in rows
        ]

    def anomaly_candidates(self, months: int | None = None) -> dict[str, Any]:
        """Выбросы, дубли и тренды считаются в Python — одинаково для
        SQLite и PostgreSQL, без диалектных функций."""
        months = months or Config.ANOMALY_MONTHS
        since = date.today() - timedelta(days=31 * months)
        rows = self.db.query(
            f"""
            SELECT t.id,
                   COALESCE(a.name, '{NO_ARTIST}') AS artist,
                   t.kind, t.amount_cents, t.currency,
                   COALESCE(t.category, '{NO_CATEGORY}') AS category,
                   t.description, t.occurred_on, t.artist_id
            FROM transactions t
            LEFT JOIN artists a ON a.id = t.artist_id
            WHERE t.occurred_on >= ?
            ORDER BY t.occurred_on
            """,
            (self.db.d(since),),
        )

        groups: dict[tuple[str, str], list[dict]] = {}
        monthly: dict[str, dict[str, Any]] = {}
        dupe_counter: dict[tuple[Any, str, int, str], int] = {}
        orphans = 0

        for r in rows:
            occurred = as_date(r["occurred_on"])
            groups.setdefault((r["kind"], r["category"]), []).append(r)

            ym = occurred.strftime("%Y-%m")
            bucket = monthly.setdefault(
                ym,
                {"ym": ym, "revenue": Decimal(0), "expenses": Decimal(0), "tx_count": 0},
            )
            amount = from_cents(r["amount_cents"])
            if r["kind"] == "income":
                bucket["revenue"] += amount
            else:
                bucket["expenses"] += amount
            bucket["tx_count"] += 1

            key = (
                r["artist_id"],
                r["kind"],
                int(r["amount_cents"]),
                occurred.isoformat(),
            )
            dupe_counter[key] = dupe_counter.get(key, 0) + 1

            if r["artist_id"] is None:
                orphans += 1

        sigma = float(Config.ANOMALY_SIGMA)
        outliers: list[dict] = []
        for (kind, category), items in groups.items():
            if len(items) < 3:
                continue
            amounts = [int(i["amount_cents"]) for i in items]
            mean = statistics.fmean(amounts)
            sd = statistics.pstdev(amounts)
            if sd <= 0:
                continue
            for item in items:
                deviation = abs(int(item["amount_cents"]) - mean)
                if deviation > sigma * sd:
                    outliers.append(
                        {
                            "id": item["id"],
                            "artist": item["artist"],
                            "kind": kind,
                            "category": category,
                            "amount": from_cents(item["amount_cents"]),
                            "currency": item["currency"],
                            "description": item["description"],
                            "occurred_on": as_date(item["occurred_on"]),
                            "avg_amount": from_cents(round(mean)),
                            "sd_amount": from_cents(round(sd)),
                            "deviation_sigmas": round(deviation / sd, 2),
                        }
                    )
        outliers.sort(key=lambda x: x["deviation_sigmas"], reverse=True)

        duplicates = [
            {
                "artist_id": k[0],
                "kind": k[1],
                "amount": from_cents(k[2]),
                "occurred_on": k[3],
                "count": v,
            }
            for k, v in sorted(dupe_counter.items(), key=lambda kv: kv[1], reverse=True)
            if v > 1
        ][:20]

        return {
            "outliers": outliers[:40],
            "duplicates": duplicates,
            "monthly": [monthly[k] for k in sorted(monthly)],
            "integrity": {"orphan_tx": orphans, "total_tx": len(rows)},
            "sigma": str(Config.ANOMALY_SIGMA),
            "period_from": since,
        }

    # ---------- выплаты ----------
    def upsert_payments_for_month(self, year: int, month: int) -> list[dict]:
        start, end = self.month_bounds(year, month)
        agg_rows = self.db.query(
            """
            SELECT artist_id,
                   SUM(CASE WHEN kind = 'income'  THEN amount_cents ELSE 0 END) AS revenue_cents,
                   SUM(CASE WHEN kind = 'expense' THEN amount_cents ELSE 0 END) AS expenses_cents
            FROM transactions
            WHERE occurred_on >= ? AND occurred_on < ? AND artist_id IS NOT NULL
            GROUP BY artist_id
            """,
            (self.db.d(start), self.db.d(end)),
        )
        agg = {int(r["artist_id"]): r for r in agg_rows}
        period_end = end - timedelta(days=1)
        created: list[dict] = []

        for artist in self.list_artists():
            data = agg.get(int(artist["id"]), {})
            revenue_cents = int(data.get("revenue_cents") or 0)
            expenses_cents = int(data.get("expenses_cents") or 0)
            rate = Decimal(str(artist["rate"] or 0))
            gross = max(
                from_cents(revenue_cents) - from_cents(expenses_cents), Decimal(0)
            )
            amount_cents = to_cents(
                (gross * rate).quantize(CENTS, rounding=ROUND_HALF_UP)
            )
            self.db.execute(
                """
                INSERT INTO payments
                    (artist_id, period_start, period_end, revenue_cents,
                     expenses_cents, rate, amount_cents, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, 'pending')
                ON CONFLICT (artist_id, period_start, period_end) DO UPDATE SET
                    revenue_cents = excluded.revenue_cents,
                    expenses_cents = excluded.expenses_cents,
                    rate = excluded.rate,
                    amount_cents = excluded.amount_cents
                WHERE payments.status = 'pending'
                """,
                (
                    artist["id"],
                    self.db.d(start),
                    self.db.d(period_end),
                    revenue_cents,
                    expenses_cents,
                    float(rate),
                    amount_cents,
                ),
            )
            created.append(
                {"artist": artist["name"], "amount": from_cents(amount_cents)}
            )
        return created

    def list_payments(self, limit: int = 20, status: str | None = None) -> list[dict]:
        sql = """
            SELECT p.id, a.name AS artist, p.period_start, p.period_end,
                   p.revenue_cents, p.expenses_cents, p.rate, p.amount_cents,
                   p.status, p.paid_at
            FROM payments p
            JOIN artists a ON a.id = p.artist_id
        """
        params: list[Any] = []
        if status:
            sql += " WHERE p.status = ?"
            params.append(status)
        sql += " ORDER BY p.period_start DESC, p.amount_cents DESC LIMIT ?"
        params.append(limit)

        rows = self.db.query(sql, params)
        return [
            {
                "id": r["id"],
                "artist": r["artist"],
                "period_start": as_date(r["period_start"]),
                "period_end": as_date(r["period_end"]),
                "revenue": from_cents(r["revenue_cents"]),
                "expenses": from_cents(r["expenses_cents"]),
                "rate": Decimal(str(r["rate"] or 0)),
                "amount": from_cents(r["amount_cents"]),
                "status": r["status"],
            }
            for r in rows
        ]

    def mark_payment_paid(self, payment_id: int) -> None:
        self.db.execute(
            "UPDATE payments SET status = 'paid', paid_at = ? WHERE id = ?",
            (self.db.now(), payment_id),
        )

    # ---------- служебное ----------
    def save_report(
        self,
        kind: str,
        body: str,
        payload: dict,
        period: tuple[date, date] | None = None,
    ) -> None:
        self.db.execute(
            """
            INSERT INTO ai_reports (kind, period_start, period_end, payload, body)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                kind,
                self.db.d(period[0]) if period else None,
                self.db.d(period[1]) if period else None,
                json_dumps(payload),
                body,
            ),
        )

    # ---------- аналитика и работа с операциями ----------
    def totals(self, start: date, end: date) -> dict:
        row = self.db.query(
            """
            SELECT
                COALESCE(SUM(CASE WHEN kind = 'income'  THEN amount_cents END), 0) AS rev,
                COALESCE(SUM(CASE WHEN kind = 'expense' THEN amount_cents END), 0) AS exp,
                COUNT(*) AS cnt
            FROM transactions
            WHERE occurred_on >= ? AND occurred_on <= ?
            """,
            (self.db.d(start), self.db.d(end)),
        )[0]
        revenue = from_cents(row["rev"] or 0)
        expenses = from_cents(row["exp"] or 0)
        return {
            "revenue": revenue,
            "expenses": expenses,
            "net": revenue - expenses,
            "count": int(row["cnt"] or 0),
        }

    def recent_transactions(self, limit: int = 10) -> list[dict]:
        rows = self.db.query(
            """
            SELECT t.id, t.kind, t.amount_cents, t.currency, t.category,
                   t.description, t.occurred_on, t.source, a.name AS artist
              FROM transactions t
              LEFT JOIN artists a ON a.id = t.artist_id
             ORDER BY t.id DESC
             LIMIT ?
            """,
            (int(limit),),
        )
        for r in rows:
            r["amount"] = from_cents(r["amount_cents"])
        return rows

    def search_transactions(self, term: str, limit: int = 15) -> list[dict]:
        """Поиск по категории, описанию, артисту, сумме и дате.

        Фильтруем в Python: SQLite умеет LOWER() только для ASCII,
        а нам нужен регистронезависимый поиск по-русски.
        """
        needle = (term or "").strip().lower()
        if not needle:
            return []
        rows = self.db.query(
            """
            SELECT t.id, t.kind, t.amount_cents, t.currency, t.category,
                   t.description, t.occurred_on, t.source, a.name AS artist
              FROM transactions t
              LEFT JOIN artists a ON a.id = t.artist_id
             ORDER BY t.occurred_on DESC, t.id DESC
             LIMIT 5000
            """,
            (),
        )
        digits = needle.replace(" ", "").replace(",", ".")
        out: list[dict] = []
        for r in rows:
            r["amount"] = from_cents(r["amount_cents"])
            haystack = " ".join(
                str(value or "")
                for value in (
                    r.get("category"),
                    r.get("description"),
                    r.get("artist"),
                    r.get("source"),
                    str(r["occurred_on"])[:10],
                    f"{r['amount']}",
                )
            ).lower()
            if needle in haystack or (digits and digits in haystack.replace(" ", "")):
                out.append(r)
                if len(out) >= int(limit):
                    break
        return out

    def period_transactions(self, start: date, end: date) -> list[dict]:
        rows = self.db.query(
            """
            SELECT t.id, t.occurred_on, t.kind, t.amount_cents, t.currency,
                   t.category, t.description, t.source, a.name AS artist
              FROM transactions t
              LEFT JOIN artists a ON a.id = t.artist_id
             WHERE t.occurred_on >= ? AND t.occurred_on <= ?
             ORDER BY t.occurred_on, t.id
            """,
            (self.db.d(start), self.db.d(end)),
        )
        for r in rows:
            r["amount"] = from_cents(r["amount_cents"])
        return rows

    def last_transaction_by(self, user_id: int) -> dict | None:
        rows = self.db.query(
            """
            SELECT t.id, t.kind, t.amount_cents, t.category, t.description,
                   t.occurred_on, a.name AS artist
              FROM transactions t
              LEFT JOIN artists a ON a.id = t.artist_id
             WHERE t.created_by = ?
             ORDER BY t.id DESC
             LIMIT 1
            """,
            (user_id,),
        )
        if not rows:
            return None
        rows[0]["amount"] = from_cents(rows[0]["amount_cents"])
        return rows[0]

    def delete_transaction(self, tx_id: int) -> int:
        return self.db.execute("DELETE FROM transactions WHERE id = ?", (int(tx_id),))

    # ---------- доступ: пользователи ----------
    def users_count(self) -> int:
        return int(self.db.scalar("SELECT COUNT(*) FROM bot_users") or 0)

    def owners_count(self) -> int:
        return int(
            self.db.scalar(
                "SELECT COUNT(*) FROM bot_users WHERE role = 'owner' AND is_active = 1"
            )
            or 0
        )

    def get_user(self, tg_user_id: int) -> dict | None:
        rows = self.db.query(
            "SELECT * FROM bot_users WHERE tg_user_id = ?", (tg_user_id,)
        )
        return rows[0] if rows else None

    def list_users(self) -> list[dict]:
        return self.db.query(
            """
            SELECT * FROM bot_users
            ORDER BY CASE role
                        WHEN 'owner' THEN 0
                        WHEN 'admin' THEN 1
                        ELSE 2
                     END,
                     tg_user_id
            """
        )

    def grant(
        self,
        tg_user_id: int,
        role: str,
        invited_by: int | None = None,
        username: str | None = None,
        full_name: str | None = None,
    ) -> dict | None:
        self.db.execute(
            """
            INSERT INTO bot_users
                (tg_user_id, username, full_name, role, is_active, invited_by)
            VALUES (?, ?, ?, ?, 1, ?)
            ON CONFLICT (tg_user_id) DO UPDATE SET
                role = excluded.role,
                is_active = 1,
                username = COALESCE(excluded.username, bot_users.username),
                full_name = COALESCE(excluded.full_name, bot_users.full_name)
            """,
            (tg_user_id, username, full_name, role, invited_by),
        )
        return self.get_user(tg_user_id)

    def set_role(self, tg_user_id: int, role: str) -> None:
        self.db.execute(
            "UPDATE bot_users SET role = ? WHERE tg_user_id = ?", (role, tg_user_id)
        )

    def set_active(self, tg_user_id: int, active: bool) -> None:
        self.db.execute(
            "UPDATE bot_users SET is_active = ? WHERE tg_user_id = ?",
            (1 if active else 0, tg_user_id),
        )

    def delete_user(self, tg_user_id: int) -> None:
        self.db.execute("DELETE FROM bot_users WHERE tg_user_id = ?", (tg_user_id,))

    def touch_user(
        self,
        tg_user_id: int,
        username: str | None = None,
        full_name: str | None = None,
    ) -> dict | None:
        """Возвращает запись доступа и обновляет профиль, либо None если доступа нет."""
        row = self.get_user(tg_user_id)
        if row is None:
            return None
        self.db.execute(
            """
            UPDATE bot_users
               SET username = COALESCE(?, username),
                   full_name = COALESCE(?, full_name),
                   last_seen_at = ?
             WHERE tg_user_id = ?
            """,
            (
                username,
                full_name,
                datetime.now().isoformat(timespec="seconds"),
                tg_user_id,
            ),
        )
        row["username"] = username or row["username"]
        row["full_name"] = full_name or row["full_name"]
        return row

    def ensure_owners(self, ids: set[int]) -> None:
        """Владельцы из OWNER_IDS всегда существуют и всегда владельцы."""
        for uid in sorted(ids):
            row = self.get_user(uid)
            if row is None:
                self.grant(uid, "owner")
            elif row["role"] != "owner" or not int(row["is_active"]):
                self.db.execute(
                    "UPDATE bot_users SET role = 'owner', is_active = 1 "
                    "WHERE tg_user_id = ?",
                    (uid,),
                )

    # ---------- доступ: приглашения ----------
    @staticmethod
    def new_code() -> str:
        alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"  # без похожих символов
        return "".join(secrets.choice(alphabet) for _ in range(8))

    def create_invite(
        self, role: str, uses: int, ttl_hours: int, created_by: int | None
    ) -> dict:
        uses = max(1, int(uses))
        expires = (
            (datetime.now() + timedelta(hours=ttl_hours)).isoformat(timespec="seconds")
            if ttl_hours > 0
            else None
        )
        for _ in range(5):  # на случай коллизии кода
            code = self.new_code()
            if not self.db.query("SELECT code FROM invites WHERE code = ?", (code,)):
                break
        self.db.execute(
            """
            INSERT INTO invites (code, role, uses_left, expires_at, created_by)
            VALUES (?, ?, ?, ?, ?)
            """,
            (code, role, uses, expires, created_by),
        )
        return {
            "code": code,
            "role": role,
            "uses_left": uses,
            "expires_at": expires,
        }

    def list_invites(self, only_active: bool = True) -> list[dict]:
        rows = self.db.query("SELECT * FROM invites ORDER BY created_at DESC")
        if not only_active:
            return rows
        now_iso = datetime.now().isoformat(timespec="seconds")
        active = []
        for r in rows:
            if int(r["is_revoked"]) or int(r["uses_left"]) <= 0:
                continue
            if r["expires_at"] and str(r["expires_at"]) < now_iso:
                continue
            active.append(r)
        return active

    def revoke_invite(self, code: str) -> None:
        self.db.execute(
            "UPDATE invites SET is_revoked = 1 WHERE code = ?", ((code or "").upper(),)
        )

    def redeem_invite(
        self,
        code: str,
        tg_user_id: int,
        username: str | None = None,
        full_name: str | None = None,
    ) -> dict | None:
        code = (code or "").strip().upper()
        if not code:
            return None
        rows = self.db.query("SELECT * FROM invites WHERE code = ?", (code,))
        if not rows:
            return None
        inv = rows[0]
        if int(inv["is_revoked"]) or int(inv["uses_left"]) <= 0:
            return None
        if inv["expires_at"] and str(inv["expires_at"]) < datetime.now().isoformat(
            timespec="seconds"
        ):
            return None
        used = self.db.execute(
            "UPDATE invites SET uses_left = uses_left - 1 "
            "WHERE code = ? AND uses_left > 0 AND is_revoked = 0",
            (code,),
        )
        if not used:
            return None
        user = self.grant(
            tg_user_id, inv["role"], inv["created_by"], username, full_name
        )
        self.audit(
            tg_user_id, "access_invite_redeem", {"code": code, "role": inv["role"]}
        )
        return user

    def audit(self, user_id: int | None, action: str, details: dict) -> None:
        self.db.execute(
            "INSERT INTO audit_log (tg_user_id, action, details) VALUES (?, ?, ?)",
            (user_id, action, json_dumps(details)),
        )


# -----------------------------------------------------------------------------
# ИИ: промпты, модели, маршрутизация
# -----------------------------------------------------------------------------
SYSTEM_ANALYST = (
    f"Ты главный бухгалтер лейбла {Config.LABEL_NAME}. Только переданные данные, "
    "без выдумок.\n"
    "Каждая сумма — с расчётом: <code>(150000 - 45000) x 0.3 = 31500</code>.\n"
    "Цифры строго из блоков РАСЧЁТЫ и ДАННЫЕ; нет числа — скажи прямо.\n"
    "Доля артиста = (Выручка - Расходы) x Ставка, остаток — прибыль лейбла.\n"
    "Пиши по-русски, кратко, тегами Telegram (<b>, <i>, <code>), без markdown, "
    "без вступлений и извинений."
)

# Добавляется к каждому текстовому запросу: без расшифровки ответ не принимается.
NUMBERS_RULE = (
    "\n\nКаждую сумму сопровождай расчётом в <code>...</code> с реальными числами. "
    "Ничего не добавляй от себя. Без воды."
)

EXCEL_PARSE_PROMPT = """Строки из таблицы лейбла (1-я строка обычно заголовки).
Верни ТОЛЬКО JSON без markdown:
{"rows":[{"artist":str|null,"kind":"income"|"expense","amount":num,"currency":"RUB","category":str|null,"description":str|null,"occurred_on":"YYYY-MM-DD"}],"errors":[str],"totals":{"income":num,"expense":num,"rows":int},"confidence":0..1,"summary":"1 предложение"}
Правила: минус в доходе => expense и amount=abs; реклама/студия/мастеринг/дистрибуция/аванс => expense; роялти/стриминг/концерт/синхро/мерч => income; неполная дата => 1-е число месяца; строки Итого/Total => в errors; сомнительное — в errors, не выбрасывать молча; структура непонятна => низкий confidence."""


# -----------------------------------------------------------------------------
# Каталог моделей и маршрутизация
# -----------------------------------------------------------------------------
TIER_BASE, TIER_CHEAP, TIER_PREMIUM, TIER_EXTREME = "base", "cheap", "premium", "extreme"

TIER_TITLES = {
    TIER_BASE: "базовая",
    TIER_CHEAP: "дешёвый помощник",
    TIER_PREMIUM: "тяжёлаяартиллерия",
    TIER_EXTREME: "крайний случай",
}

# Порядок в словаре = порядок в меню выбора модели.
MODEL_CATALOG: dict[str, dict[str, Any]] = {
    "qwen": {
        "title": "Qwen 3.8 Max",
        "api": Config.MODEL_QWEN,
        "tier": TIER_BASE,
        "price": "₽",
        "best": "ежедневная работа: ввод операций, Excel, быстрые вопросы",
        "about": (
            "Модель по умолчанию. Быстрая и дешёвая, аккуратно разбирает "
            "таблицы и свободный текст. Если сама не уверена — передаёт задачу дальше."
        ),
    },
    "sonnet5": {
        "title": "Sonnet 5",
        "api": Config.MODEL_SONNET5,
        "tier": TIER_CHEAP,
        "price": "₽₽",
        "best": "проверка расчётов и аккуратные тексты отчётов",
        "about": (
            "Главный дешёвый помощник. Лучше Qwen в формулировках и внимательнее "
            "к деталям расчёта. Включается первым, когда Qwen сомневается."
        ),
    },
    "gpt55": {
        "title": "GPT 5.5",
        "api": Config.MODEL_GPT55,
        "tier": TIER_CHEAP,
        "price": "₽₽",
        "best": "вторая пара глаз по арифметике и короткие сверки",
        "about": (
            "Второй дешёвый помощник. Крепкая логика и счёт, хорошо ловит "
            "арифметические ошибки в чужих выводах."
        ),
    },
    "opus5": {
        "title": "Opus 5",
        "api": Config.MODEL_OPUS5,
        "tier": TIER_PREMIUM,
        "price": "₽₽₽₽",
        "best": "аудит, поиск скрытых ошибок, сложные аномалии",
        "about": (
            "Самая внимательная модель по цифрам. Берётся, когда надо понять, "
            "почему не сходятся итоги или кто из артистов реально убыточен."
        ),
    },
    "luna": {
        "title": "GPT 5.6 Luna",
        "api": Config.MODEL_LUNA,
        "tier": TIER_PREMIUM,
        "price": "₽₽₽₽",
        "best": "большие разборы: много артистов, спорные месяцы, стратегия",
        "about": (
            "Держит в голове много данных сразу и пишет развёрнутые разборы "
            "с выводами. Подходит для квартальных и годовых итогов."
        ),
    },
    "terra": {
        "title": "GPT 5.6 Terra",
        "api": Config.MODEL_TERRA,
        "tier": TIER_EXTREME,
        "price": "₽₽₽₽₽",
        "best": "расследование расхождений на крупные суммы",
        "about": (
            "Дорогая и медленная. Включается автоматически только в крайних "
            "случаях, когда остальные модели не сошлись в цифрах."
        ),
    },
    "sol": {
        "title": "GPT 5.6 Sol",
        "api": Config.MODEL_SOL,
        "tier": TIER_EXTREME,
        "price": "₽₽₽₽₽",
        "best": "самые тяжёлые задачи: пересчёт года, налоговые сценарии",
        "about": (
            "Потолок качества и цены. Вызывается крайне редко — когда цена "
            "ошибки выше цены запроса."
        ),
    },
}

CHEAP_HELPERS = ("sonnet5", "gpt55")
PREMIUM_MODELS = ("opus5", "luna")
EXTREME_MODELS = ("terra", "sol")

# Сложность задач: 1 — мелочь, 4 — только тяжёлая артиллерия.
TASK_COMPLEXITY = {
    "quick": 1,
    "excel": 2,
    "ask": 2,
    "report": 3,
    "anomaly": 3,
    "audit": 4,
}


def normalize_model(key: str | None) -> str | None:
    """Понимает и короткие ключи, и человеческие названия моделей."""
    if not key:
        return None
    low = str(key).strip().lower()
    if low in ("", "auto", "авто"):
        return None
    if low in MODEL_CATALOG:
        return low
    for mkey, spec in MODEL_CATALOG.items():
        if low in (spec["title"].lower(), str(spec["api"]).lower()):
            return mkey
    aliases = {
        "qwen3.8": "qwen",
        "qwen-3.8-max": "qwen",
        "sonnet": "sonnet5",
        "opus": "opus5",
        "gpt5.5": "gpt55",
        "gpt-5.5": "gpt55",
        "gpt5.6-luna": "luna",
        "gpt5.6-terra": "terra",
        "gpt5.6-sol": "sol",
    }
    return aliases.get(low)


def base_model() -> str:
    return normalize_model(Config.AI_BASE_MODEL) or "qwen"


def route_chain(task: str, manual: str | None = None, hard: bool = False) -> list[str]:
    """Цепочка моделей: кто начинает и кому передаёт, если не уверен.

    Ручной выбор — ровно одна модель, без эскалации.
    Авто: Qwen -> дешёвые помощники -> тяжёлые -> крайний случай.
    """
    manual_key = normalize_model(manual)
    if manual_key:
        return [manual_key]

    complexity = int(TASK_COMPLEXITY.get(task, 2))
    if hard:
        complexity = max(complexity, 3)

    chain = [base_model()]
    # шаг 2: дешёвые помощники — сначала тот, кто лучше для этого типа задач
    if task in ("quick", "excel"):
        chain += ["gpt55", "sonnet5"]
    else:
        chain += ["sonnet5", "gpt55"]
    # шаг 3: тяжёлая артиллерия — только для сложных задач
    if complexity >= 3:
        chain += ["opus5", "luna"] if task in ("anomaly", "audit") else ["luna", "opus5"]
    # шаг 4: Terra и Sol — только крайние случаи и только если разрешено
    if complexity >= 4 and Config.AI_ALLOW_EXTREME:
        chain += list(EXTREME_MODELS)

    seen: set[str] = set()
    out: list[str] = []
    for key in chain:
        if key in MODEL_CATALOG and key not in seen:
            seen.add(key)
            out.append(key)
    return out


def models_help() -> str:
    """Описание всех моделей для меню выбора."""
    lines = []
    for spec in MODEL_CATALOG.values():
        lines.append(
            f"<b>{spec['title']}</b> {spec['price']} · <i>{TIER_TITLES[spec['tier']]}</i>\n"
            f"Лучше всего: {spec['best']}\n{spec['about']}"
        )
    return "\n\n".join(lines)


# -----------------------------------------------------------------------------
# Проверка цифр в ответе модели
# -----------------------------------------------------------------------------
NUM_TOKEN_RE = re.compile(r"\d[\d\s\u00a0\u202f.,]*\d|\d")


def _norm_num(token: str) -> Decimal | None:
    """«150 000,50» / «150000.50» / «150,000.50» -> Decimal."""
    raw = token.strip().replace("\u00a0", "").replace("\u202f", "").replace(" ", "")
    if not raw:
        return None
    if "," in raw and "." in raw:
        raw = (
            raw.replace(",", "")
            if raw.rfind(".") > raw.rfind(",")
            else raw.replace(".", "").replace(",", ".")
        )
    elif raw.count(",") == 1 and len(raw.split(",")[1]) in (1, 2):
        raw = raw.replace(",", ".")
    else:
        raw = raw.replace(",", "")
    if raw.count(".") > 1:
        raw = raw.replace(".", "")
    try:
        return Decimal(raw).quantize(CENTS)
    except (InvalidOperation, ValueError):
        return None


def collect_numbers(obj: Any, acc: set[Decimal] | None = None) -> set[Decimal]:
    """Все числа, которые реально есть в данных (белый список)."""
    acc = acc if acc is not None else set()
    if isinstance(obj, dict):
        for value in obj.values():
            collect_numbers(value, acc)
    elif isinstance(obj, (list, tuple, set)):
        for value in obj:
            collect_numbers(value, acc)
    elif isinstance(obj, bool):
        return acc
    elif isinstance(obj, (int, float, Decimal)):
        num = _norm_num(str(obj))
        if num is not None:
            acc.add(num)
    elif isinstance(obj, str):
        for token in NUM_TOKEN_RE.findall(obj):
            num = _norm_num(token)
            if num is not None:
                acc.add(num)
    return acc


def _matches_allowed(num: Decimal, allowed: set[Decimal]) -> bool:
    for value in allowed:
        if value == num:
            return True
        gap = abs(value - num)
        if gap <= max(Decimal("1"), abs(value) * Decimal("0.01")):
            return True
        # модель могла округлить до тысяч или до целых
        for step in (Decimal("1000"), Decimal("100"), Decimal("1")):
            if (value / step).quantize(Decimal("1")) == (num / step).quantize(
                Decimal("1")
            ):
                return True
    return False


def verify_numbers(
    text: str, allowed: set[Decimal], floor: Decimal = Decimal("1000"), limit: int = 6
) -> list[str]:
    """Возвращает крупные числа из ответа, которых нет в исходных данных."""
    if not allowed:
        return []
    suspicious: list[str] = []
    seen: set[Decimal] = set()
    for token in NUM_TOKEN_RE.findall(text or ""):
        num = _norm_num(token)
        if num is None or num in seen:
            continue
        if abs(num) < floor:  # проценты, даты, счётчики не проверяем
            continue
        if 1900 <= num <= 2200 and num == num.to_integral_value():  # годы
            continue
        seen.add(num)
        if not _matches_allowed(num, allowed):
            suspicious.append(f"{num:,.2f}".replace(",", " ").replace(".", ","))
        if len(suspicious) >= limit:
            break
    return suspicious


def extract_confidence(text: str) -> Decimal | None:
    match = re.search(r'"confidence"\s*:\s*([0-9]*\.?[0-9]+)', text or "")
    if not match:
        return None
    try:
        return Decimal(match.group(1))
    except InvalidOperation:
        return None


# -----------------------------------------------------------------------------
# Провайдеры и клиент
# -----------------------------------------------------------------------------
class ProviderError(RuntimeError):
    def __init__(self, message: str, status: int | None = None) -> None:
        super().__init__(message)
        self.status = status


PROVIDER_SPECS = {
    "xkiro": {
        "title": "Xkiro",
        "key_env": "XKIRO_API_KEY",
        "url_attr": "XKIRO_BASE_URL",
        "key_attr": "XKIRO_API_KEY",
    },
    "tooken": {
        "title": "Tooken",
        "key_env": "TOOKEN_API_KEY",
        "url_attr": "TOOKEN_BASE_URL",
        "key_attr": "TOOKEN_API_KEY",
    },
}


def active_providers() -> list[dict]:
    """Провайдеры с ключом, в порядке AI_PROVIDER_ORDER."""
    order = [
        item.strip().lower()
        for item in Config.AI_PROVIDER_ORDER.split(",")
        if item.strip()
    ] or list(PROVIDER_SPECS)
    out: list[dict] = []
    for name in order:
        spec = PROVIDER_SPECS.get(name)
        if not spec:
            continue
        api_key = getattr(Config, spec["key_attr"], "")
        if not api_key:
            continue
        out.append(
            {
                "name": name,
                "title": spec["title"],
                "base_url": getattr(Config, spec["url_attr"], ""),
                "api_key": api_key,
            }
        )
    return out


# -----------------------------------------------------------------------------
# Экономия токенов: сжатые данные, счётчик расхода, кэш ответов
# -----------------------------------------------------------------------------
def json_min(obj: Any) -> str:
    """JSON без лишних пробелов — на 10-15% меньше токенов, чем обычный."""

    def default(o: Any) -> Any:
        if isinstance(o, Decimal):
            return float(o)
        if isinstance(o, (date, datetime)):
            return o.isoformat()[:10]
        return str(o)

    return json.dumps(obj, ensure_ascii=False, separators=(",", ":"), default=default)


def _short_value(value: Any) -> Any:
    """Округляем деньги до рублей и режем текст: копейки в промпте не нужны."""
    if isinstance(value, Decimal):
        return int(value.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    if isinstance(value, float):
        return round(value, 2)
    if isinstance(value, (date, datetime)):
        return str(value)[:10]
    if isinstance(value, str):
        return value[:90]
    return value


def trim_data(data: Any, rows: int | None = None, depth: int = 0) -> Any:
    """Сжимает payload перед отправкой в модель: короткие списки, без пустых ключей."""
    limit = rows or Config.AI_MAX_DATA_ROWS
    if isinstance(data, dict):
        out: dict[str, Any] = {}
        for key, value in data.items():
            if value in (None, "", [], {}):
                continue
            out[str(key)] = trim_data(value, limit, depth + 1)
        return out
    if isinstance(data, (list, tuple)):
        items = [trim_data(v, limit, depth + 1) for v in list(data)[:limit]]
        if len(data) > limit:
            items.append(f"...ещё {len(data) - limit} строк не отправлено (экономия токенов)")
        return items
    return _short_value(data)


class TokenMeter:
    """Считает расход токенов за сутки и сколько сэкономлено на кэше."""

    def __init__(self) -> None:
        self.lock = threading.Lock()
        self.day = date.today()
        self.prompt = 0
        self.completion = 0
        self.calls = 0
        self.cached = 0
        self.saved = 0
        self.by_model: dict[str, int] = {}

    def _roll(self) -> None:
        today = date.today()
        if today != self.day:
            self.day = today
            self.prompt = self.completion = self.calls = self.cached = self.saved = 0
            self.by_model = {}

    def add(self, model_key: str, usage: dict | None) -> None:
        usage = usage or {}
        with self.lock:
            self._roll()
            prompt = int(usage.get("prompt_tokens") or 0)
            completion = int(usage.get("completion_tokens") or 0)
            self.prompt += prompt
            self.completion += completion
            self.calls += 1
            self.by_model[model_key] = (
                self.by_model.get(model_key, 0) + prompt + completion
            )

    def add_cached(self, approx_tokens: int) -> None:
        with self.lock:
            self._roll()
            self.cached += 1
            self.saved += max(0, int(approx_tokens))

    def total(self) -> int:
        with self.lock:
            self._roll()
            return self.prompt + self.completion

    def over_budget(self) -> bool:
        limit = Config.AI_DAILY_TOKEN_LIMIT
        return bool(limit) and self.total() >= limit

    def stats_text(self) -> str:
        with self.lock:
            self._roll()
            total = self.prompt + self.completion
            top = sorted(self.by_model.items(), key=lambda kv: -kv[1])[:3]
            parts = [
                f"За сегодня: <b>{total}</b> токенов за {self.calls} запросов",
                f"Ввод {self.prompt} + ответ {self.completion}",
            ]
            if self.cached:
                parts.append(
                    f"Из кэша: {self.cached} раз, сэкономлено ≈{self.saved} токенов"
                )
            if Config.AI_DAILY_TOKEN_LIMIT:
                parts.append(
                    f"Лимит в сутки: {Config.AI_DAILY_TOKEN_LIMIT}"
                    + (" — исчерпан, работают только дешёвые модели" if total >= Config.AI_DAILY_TOKEN_LIMIT else "")
                )
            if top:
                parts.append(
                    "По моделям: "
                    + ", ".join(
                        f"{MODEL_CATALOG[k]['title']} {v}"
                        for k, v in top
                        if k in MODEL_CATALOG
                    )
                )
            return "\n".join(parts)


TOKENS = TokenMeter()

_AI_CACHE: dict[str, tuple[float, dict]] = {}
_AI_CACHE_LOCK = threading.Lock()


def cache_key(task: str, prompt: str, model_key: str, json_mode: bool) -> str:
    raw = f"{task}|{model_key}|{int(json_mode)}|{prompt}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def cache_get(key: str) -> dict | None:
    if not Config.AI_ECONOMY or Config.AI_CACHE_TTL <= 0:
        return None
    now = time.time()
    with _AI_CACHE_LOCK:
        hit = _AI_CACHE.get(key)
        if not hit:
            return None
        stamp, payload = hit
        if now - stamp > Config.AI_CACHE_TTL:
            _AI_CACHE.pop(key, None)
            return None
        result = dict(payload)
    result["cached"] = True
    TOKENS.add_cached(result.get("approx_tokens", 0))
    return result


def cache_put(key: str, result: dict) -> None:
    if not Config.AI_ECONOMY or Config.AI_CACHE_TTL <= 0:
        return
    with _AI_CACHE_LOCK:
        _AI_CACHE[key] = (time.time(), dict(result))
        if len(_AI_CACHE) > max(8, Config.AI_CACHE_SIZE):
            for old_key in sorted(_AI_CACHE, key=lambda k: _AI_CACHE[k][0])[:8]:
                _AI_CACHE.pop(old_key, None)


def econ_chain(chain: list[str]) -> list[str]:
    """Режет цепочку моделей: меньше шагов = меньше токенов."""
    if not chain:
        return chain
    if TOKENS.over_budget():
        cheap = [
            key
            for key in chain
            if MODEL_CATALOG[key]["tier"] in (TIER_BASE, TIER_CHEAP)
        ]
        chain = cheap or chain[:1]
        log.info("Суточный лимит токенов исчерпан: только дешёвые модели")
    if not Config.AI_ECONOMY:
        return chain
    return chain[: max(1, Config.AI_MAX_STEPS)]


class LLMService:
    """OpenAI-совместимый клиент: Xkiro + Tooken, автовыбор модели, эскалация."""

    def __init__(self) -> None:
        self.providers = active_providers()
        self.lock = threading.Lock()
        self.calls: list[dict] = []  # короткая история для экрана диагностики

    @property
    def ready(self) -> bool:
        return bool(self.providers)

    def provider_names(self) -> str:
        return ", ".join(p["title"] for p in self.providers) or "нет ключей"

    # ---------- транспорт ----------
    def _post(self, provider: dict, payload: dict) -> dict:
        import urllib.error
        import urllib.request

        url = provider["base_url"].rstrip("/") + "/chat/completions"
        request = urllib.request.Request(
            url,
            data=json_dumps(payload).encode("utf-8"),
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {provider['api_key']}",
                "User-Agent": "finbot-2217",
            },
            method="POST",
        )
        try:
            with urllib.request.urlopen(request, timeout=Config.AI_TIMEOUT) as resp:
                return json.loads(resp.read().decode("utf-8", "replace"))
        except urllib.error.HTTPError as exc:
            body = exc.read().decode("utf-8", "replace")[:300]
            raise ProviderError(
                f"{provider['title']}: HTTP {exc.code} {body}", status=exc.code
            ) from exc
        except Exception as exc:
            raise ProviderError(f"{provider['title']}: {exc}") from exc

    @staticmethod
    def _extract_text(data: dict) -> str:
        choices = data.get("choices") or []
        if not choices:
            return ""
        first = choices[0] or {}
        message = first.get("message") or {}
        content = message.get("content")
        if isinstance(content, list):  # часть провайдеров отдаёт список блоков
            content = "".join(
                str(part.get("text") or "")
                for part in content
                if isinstance(part, dict)
            )
        return str(content or first.get("text") or "").strip()

    def _call_model(
        self,
        spec: dict,
        system: str,
        prompt: str,
        max_tokens: int | None,
        json_mode: bool,
    ) -> tuple[str, str]:
        payload: dict[str, Any] = {
            "model": spec["api"],
            "messages": [
                {"role": "system", "content": system},
                {"role": "user", "content": prompt},
            ],
            "max_tokens": int(max_tokens or Config.AI_MAX_TOKENS),
            "temperature": float(Config.AI_TEMPERATURE),
        }
        if json_mode:
            payload["response_format"] = {"type": "json_object"}

        last: ProviderError | None = None
        for provider in self.providers:
            try:
                data = self._post(provider, payload)
            except ProviderError as exc:
                log.warning("Провайдер не ответил: %s", exc)
                last = exc
                continue
            text = self._extract_text(data)
            if text:
                usage = data.get("usage")
                return (
                    text,
                    provider["title"],
                    usage if isinstance(usage, dict) else {},
                )
            last = ProviderError(f"{provider['title']}: пустой ответ")
        raise last or ProviderError("нет настроенных провайдеров ИИ")

    # ---------- основной вызов ----------
    def _run(
        self,
        task: str,
        prompt: str,
        system: str | None,
        max_tokens: int | None,
        json_mode: bool,
        manual: str | None,
        hard: bool,
        allowed: set[Decimal] | None,
    ) -> dict:
        if not self.providers:
            raise RuntimeError(
                "Не задан ни один ключ ИИ. Заполните XKIRO_API_KEY или TOOKEN_API_KEY "
                "в переменных окружения."
            )

        chain = route_chain(task, manual, hard)
        if not manual:
            chain = econ_chain(chain)
        base_system = (system or SYSTEM_ANALYST) + NUMBERS_RULE
        key = cache_key(task, prompt, chain[0], json_mode)
        hit = cache_get(key)
        if hit is not None:
            log.info("Ответ из кэша без расхода токенов: %s", task)
            return hit
        problems: list[str] = []
        spent = 0
        text, used, provider_title = "", "", ""
        confidence: Decimal | None = None
        unverified: list[str] = []

        for step, model_key in enumerate(chain):
            spec = MODEL_CATALOG[model_key]
            extra = ""
            if step and problems:
                extra = (
                    "\n\nВНИМАНИЕ: предыдущая модель не справилась — "
                    f"{problems[-1]}. Пересчитай сам и покажи все формулы с числами."
                )
            try:
                text, provider_title, usage = self._call_model(
                    spec, base_system, prompt + extra, max_tokens, json_mode
                )
                TOKENS.add(model_key, usage)
                spent += int((usage or {}).get("total_tokens") or 0)
            except ProviderError as exc:
                problems.append(str(exc))
                continue

            used = model_key
            confidence = extract_confidence(text) if json_mode else None
            unverified = (
                verify_numbers(text, allowed or set()) if not json_mode else []
            )
            low_confidence = (
                confidence is not None and confidence < Config.AI_CONFIDENCE_MIN
            )
            if not low_confidence and not unverified:
                break
            if low_confidence:
                problems.append(f"низкая уверенность {confidence}")
            if unverified:
                problems.append(
                    "в ответе числа, которых нет в данных: "
                    + ", ".join(unverified)
                )
            log.info("AI эскалация с %s: %s", model_key, problems[-1])

        if not used:
            raise RuntimeError(
                "Ни одна модель не ответила. " + "; ".join(problems[-2:])
            )

        spec = MODEL_CATALOG[used]
        ACTIVE_MODEL[""] = spec["title"]
        result = {
            "text": text,
            "model": used,
            "title": spec["title"],
            "provider": provider_title,
            "path": [MODEL_CATALOG[k]["title"] for k in chain[: chain.index(used) + 1]],
            "confidence": confidence,
            "unverified": unverified,
            "task": task,
            "tokens": spent,
            "approx_tokens": spent or len(prompt) // 3,
            "cached": False,
        }
        cache_put(key, result)
        with self.lock:
            self.calls.append(
                {
                    "время": datetime.now().strftime("%d.%m %H:%M"),
                    "задача": task,
                    "модель": spec["title"],
                    "провайдер": provider_title,
                    "маршрут": " -> ".join(result["path"]),
                }
            )
            del self.calls[:-20]
        return result

    async def run(
        self,
        task: str,
        prompt: str,
        *,
        system: str | None = None,
        max_tokens: int | None = None,
        json_mode: bool = False,
        manual: str | None = None,
        hard: bool = False,
        allowed: set[Decimal] | None = None,
    ) -> dict:
        return await asyncio.to_thread(
            self._run,
            task,
            prompt,
            system,
            max_tokens,
            json_mode,
            manual,
            hard,
            allowed,
        )

    async def complete(self, prompt: str, **kw: Any) -> str:
        res = await self.run(kw.pop("task", "ask"), prompt, **kw)
        return res["text"]

    @staticmethod
    def _extract_json(text: str) -> dict:
        text = re.sub(
            r"^```(?:json)?|```$", "", (text or "").strip(), flags=re.MULTILINE
        ).strip()
        start, end = text.find("{"), text.rfind("}")
        if start == -1 or end == -1:
            raise ValueError("модель вернула ответ без JSON")
        return json.loads(text[start : end + 1])

    # ---------- прикладные задачи ----------
    async def parse_excel(self, raw_rows: list[list[Any]], manual: str | None = None) -> dict:
        rows = raw_rows[: max(5, Config.AI_MAX_EXCEL_AI_ROWS)]
        payload = json_min({"rows": rows})
        budget = min(6000, 220 + 90 * len(rows))
        res = await self.run(
            "excel",
            f"{EXCEL_PARSE_PROMPT}\n\nДанные:\n{payload}",
            max_tokens=budget,
            json_mode=True,
            manual=manual,
        )
        parsed = await asyncio.to_thread(self._extract_json, res["text"])
        parsed["_route"] = res
        return parsed

    async def quick_parse(
        self,
        text: str,
        artists: list[str],
        today: date,
        manual: str | None = None,
    ) -> dict:
        """Разбор операции, написанной человеческим языком."""
        prompt = (
            f"Разбери операцию из текста. Сегодня: {today.isoformat()}. "
            f"Артисты: {json_min(artists[:40])}.\n"
            f"Текст: {text!r}\n"
            'Верни ТОЛЬКО JSON: {"understood":bool,"kind":"income"|"expense",'
            '"amount":num,"category":str,"description":str,"artist":str|null,'
            '"occurred_on":"YYYY-MM-DD","confidence":0..1,"calc":"арифметика суммы",'
            '"note":"кратко"}\n'
            "Не операция => understood=false. Зашло/пришло/роялти/выплата => income; "
            "потратил/заплатил/реклама/студия => expense. Понимай вчера, 5 числа, "
            "прошлый месяц; без даты — сегодня. Артист только из списка, иначе null. "
            "Категория одним словом. 15к = 15000, 1,2к = 1200."
        )
        res = await self.run(
            "quick", prompt, max_tokens=400, json_mode=True, manual=manual
        )
        parsed = await asyncio.to_thread(self._extract_json, res["text"])
        parsed["_route"] = res
        return parsed

    async def ask(
        self,
        question: str,
        data: dict,
        calc_lines: list[str] | None = None,
        manual: str | None = None,
        hard: bool = False,
    ) -> dict:
        calc_block = (
            "РАСЧЁТЫ (посчитаны кодом, им можно верить):\n"
            + "\n".join(calc_lines[:14])
            if calc_lines
            else ""
        )
        prompt = (
            f"Вопрос владельца: {question}\n\n{calc_block}\n\n"
            f"ДАННЫЕ:\n{json_min(trim_data(data))}\n\n"
            "Формат: 1) одна строка с главной цифрой; "
            "2) <b>Как посчитано</b> — 2-4 формулы в <code>...</code>; "
            "3) <b>Что делать</b> — 2 пункта. До 900 символов."
        )
        allowed = collect_numbers(data)
        if calc_lines:
            collect_numbers(calc_lines, allowed)
        return await self.run(
            "ask",
            prompt,
            max_tokens=Config.AI_MAX_TOKENS,
            manual=manual,
            hard=hard,
            allowed=allowed,
        )

    async def period_report(
        self,
        label: str,
        payload: dict,
        calc_lines: list[str],
        manual: str | None = None,
    ) -> dict:
        prompt = (
            f"Отчёт лейбла за {label}.\nРАСЧЁТЫ (код):\n"
            + "\n".join(calc_lines[:18])
            + f"\n\nДАННЫЕ:\n{json_min(trim_data(payload))}\n\n"
            "Структура: 1) <b>Итоги</b> — выручка, расходы, чистыми, маржа, "
            "артистам, прибыль лейбла (каждая цифра с формулой); "
            "2) <b>Артисты</b> — топ-3 и убыточные с (Выручка - Расходы) x Ставка; "
            "3) <b>Расходы</b> — главные статьи с долей %; "
            "4) <b>Проверить руками</b> — 2-3 пункта; "
            "5) <b>Рекомендации</b> — 2 пункта с цифрами. "
            "До 1600 символов, теги Telegram."
        )
        allowed = collect_numbers(payload)
        collect_numbers(calc_lines, allowed)
        return await self.run(
            "report",
            prompt,
            max_tokens=Config.AI_MAX_TOKENS_LONG,
            manual=manual,
            allowed=allowed,
        )

    async def anomaly_report(
        self, data: dict, calc_lines: list[str] | None = None, manual: str | None = None
    ) -> dict:
        prompt = (
            "Найди аномалии и ошибки учёта.\n"
            + (
                "РАСЧЁТЫ (код):\n" + "\n".join(calc_lines[:12]) + "\n\n"
                if calc_lines
                else ""
            )
            + f"ДАННЫЕ:\n{json_min(trim_data(data))}\n\n"
            "Верни: 1) <b>КРИТИЧНО</b> — с расчётом отклонения от среднего; "
            "2) <b>ВНИМАНИЕ</b> — тренды с цифрами; "
            "3) <b>ОШИБКИ ДАННЫХ</b> — дубли, без артиста, нереальные значения; "
            "4) <b>Чек-лист</b> — 3 пункта. Аномалий нет — скажи прямо. "
            "До 1200 символов, теги Telegram."
        )
        allowed = collect_numbers(data)
        if calc_lines:
            collect_numbers(calc_lines, allowed)
        return await self.run(
            "anomaly",
            prompt,
            max_tokens=Config.AI_MAX_TOKENS,
            manual=manual,
            allowed=allowed,
        )


def route_footer(res: dict) -> str:
    """Подпись под ответом: кто считал и была ли эскалация."""
    path = res.get("path") or []
    chain = " → ".join(path) if len(path) > 1 else res.get("title", "")
    tail = f"\n\n<i>\U0001f9e0 {chain} · {res.get('provider', '')}"
    if res.get("cached"):
        tail += " · из кэша, 0 токенов"
    elif res.get("tokens"):
        tail += f" · {res['tokens']} токенов"
    tail += "</i>"
    if res.get("unverified"):
        tail += (
            "\n⚠️ <i>В ответе есть числа, которых нет в базе: "
            + ", ".join(res["unverified"])
            + ". Проверьте вручную через «\U0001f9fe Как посчитано».</i>"
        )
    return tail


# -----------------------------------------------------------------------------
# Форматирование и парсинг
# -----------------------------------------------------------------------------
def money(value: Any, currency: str = "") -> str:
    d = Decimal(str(value or 0)).quantize(CENTS)
    s = f"{d:,.2f}".replace(",", " ").replace(".", ",")
    return f"{s} {currency}".strip()


def parse_amount(text: str) -> Decimal:
    cleaned = re.sub(r"[^\d,.\-]", "", text or "").replace(",", ".")
    if cleaned.count(".") > 1:  # 1.234.567 -> 1234567
        head, _, tail = cleaned.rpartition(".")
        cleaned = head.replace(".", "") + "." + tail
    try:
        value = Decimal(cleaned).quantize(CENTS, rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(
            "Не смог распознать сумму. Пример: 125000 или 125 000,50"
        ) from exc
    if value <= 0:
        raise ValueError("Сумма должна быть больше нуля.")
    return value


def parse_rate(text: str) -> Decimal:
    raw = parse_amount(text)
    rate = raw / Decimal(100) if raw > 1 else raw
    if not (Decimal(0) < rate <= Decimal(1)):
        raise ValueError(
            "Ставка должна быть в диапазоне 1-100% (например 20 или 0.2)."
        )
    return rate.quantize(Decimal("0.0001"))


def parse_date(text: str) -> date:
    text = (text or "").strip().lower()
    if text in {"", "сегодня", "-"}:
        return date.today()
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError("Не смог распознать дату. Формат: ДД.ММ.ГГГГ или «сегодня».")


def chunks(text: str, size: int = 3800) -> list[str]:
    return [text[i : i + size] for i in range(0, len(text), size)] or ["—"]

# -----------------------------------------------------------------------------
# Любые периоды: любой день, месяц, квартал или год
# Всюду используется пара (start, end), где end НЕ включается.
# -----------------------------------------------------------------------------
MONTHS_RU = (
    "январь",
    "февраль",
    "март",
    "апрель",
    "май",
    "июнь",
    "июль",
    "август",
    "сентябрь",
    "октябрь",
    "ноябрь",
    "декабрь",
)

MONTH_ALIASES = {
    "янв": 1, "фев": 2, "мар": 3, "апр": 4, "май": 5, "мая": 5,
    "июн": 6, "июл": 7, "авг": 8, "сен": 9, "окт": 10, "ноя": 11, "дек": 12,
}


def month_name_ru(month: int) -> str:
    return MONTHS_RU[(int(month) - 1) % 12]


def add_months(base: date, count: int) -> date:
    total = (base.year * 12 + base.month - 1) + count
    return date(total // 12, total % 12 + 1, 1)


def month_period(year: int, month: int) -> tuple[date, date]:
    start = date(year, month, 1)
    return start, add_months(start, 1)


def period_label(start: date, end: date) -> str:
    """Человеческое название периода — чтобы не путаться в датах."""
    last = end - timedelta(days=1)
    if last < start:
        last = start
    if start == last:
        return start.strftime("%d.%m.%Y")
    if start.day == 1 and end == add_months(start, 1):
        return f"{month_name_ru(start.month)} {start.year}"
    if start == date(start.year, 1, 1) and end == date(start.year + 1, 1, 1):
        return f"{start.year} год"
    if start.day == 1 and end == add_months(start, 3) and start.month in (1, 4, 7, 10):
        return f"{(start.month - 1) // 3 + 1} квартал {start.year}"
    return f"{start.strftime('%d.%m.%Y')} — {last.strftime('%d.%m.%Y')}"


def parse_period(text: str, today: date | None = None) -> tuple[date, date] | None:
    """Понимает почти любой способ задать период руками.

    Поддерживает: «сегодня», «вчера», «7 дней», «этот месяц», «прошлый месяц»,
    «квартал», «q3 2026», «2026», «08.2026», «август 2026», «15.08.2026»,
    «01.01.2026-31.03.2026».
    """
    today = today or date.today()
    raw = (text or "").strip().lower().replace(" ", " ")
    if not raw:
        return None
    raw = re.sub(r"\s+", " ", raw)

    if raw in ("сегодня", "today"):
        return today, today + timedelta(days=1)
    if raw in ("вчера", "yesterday"):
        return today - timedelta(days=1), today
    if raw in ("неделя", "за неделю", "7 дней", "последние 7 дней"):
        return today - timedelta(days=6), today + timedelta(days=1)
    if raw in ("30 дней", "последние 30 дней", "месяц назад"):
        return today - timedelta(days=29), today + timedelta(days=1)
    if raw in ("месяц", "этот месяц", "текущий месяц"):
        return month_period(today.year, today.month)
    if raw in ("прошлый месяц", "предыдущий месяц"):
        first = date(today.year, today.month, 1)
        return add_months(first, -1), first
    if raw in ("квартал", "этот квартал"):
        start = date(today.year, (today.month - 1) // 3 * 3 + 1, 1)
        return start, add_months(start, 3)
    if raw in ("год", "этот год", "текущий год"):
        return date(today.year, 1, 1), date(today.year + 1, 1, 1)
    if raw in ("прошлый год",):
        return date(today.year - 1, 1, 1), date(today.year, 1, 1)
    if raw in ("всё время", "все время", "вся история"):
        return date(2000, 1, 1), today + timedelta(days=1)

    m = re.fullmatch(r"(\d+)\s*дн(?:ей|я)?", raw)
    if m:
        days = max(1, min(int(m.group(1)), 3650))
        return today - timedelta(days=days - 1), today + timedelta(days=1)

    # диапазон: 01.01.2026-31.03.2026 или 01.01.2026 — 31.03.2026
    parts = re.split(r"\s*(?:-|—|–|по|\.\.)\s*", raw)
    if len(parts) == 2 and all(re.search(r"\d{1,2}[.\-/]\d{1,2}", part) for part in parts):
        try:
            first, second = parse_date(parts[0]), parse_date(parts[1])
        except ValueError:
            first = second = None
        if first and second:
            if second < first:
                first, second = second, first
            return first, second + timedelta(days=1)

    m = re.fullmatch(r"(?:q|кв)\s*([1-4])\s*(\d{4})", raw) or re.fullmatch(
        r"([1-4])\s*кв(?:артал)?\s*(\d{4})", raw
    )
    if m:
        start = date(int(m.group(2)), (int(m.group(1)) - 1) * 3 + 1, 1)
        return start, add_months(start, 3)

    m = re.fullmatch(r"(\d{4})", raw)
    if m and 2000 <= int(m.group(1)) <= 2100:
        year = int(m.group(1))
        return date(year, 1, 1), date(year + 1, 1, 1)

    m = re.fullmatch(r"(\d{1,2})[.\-/](\d{4})", raw)
    if m and 1 <= int(m.group(1)) <= 12:
        return month_period(int(m.group(2)), int(m.group(1)))

    m = re.fullmatch(r"([а-яё]{3,})\s*(\d{4})?", raw)
    if m:
        month = MONTH_ALIASES.get(m.group(1)[:3])
        if month:
            year = int(m.group(2)) if m.group(2) else today.year
            return month_period(year, month)

    try:  # конкретный день
        day = parse_date(raw)
    except ValueError:
        return None
    return day, day + timedelta(days=1)


def kb_period(action: str, extra_rows: list | None = None) -> InlineKeyboardMarkup:
    """Единый выбор периода для сводки, отчёта, экспорта и анализа."""
    today = date.today()
    first = date(today.year, today.month, 1)
    prev_first = add_months(first, -1)
    quarter = date(today.year, (today.month - 1) // 3 * 3 + 1, 1)

    def item(title: str, start: date, end: date) -> InlineKeyboardButton:
        return InlineKeyboardButton(
            title, callback_data=f"period:go:{action}:{start.isoformat()}:{end.isoformat()}"
        )

    rows = [
        [
            item("Сегодня", today, today + timedelta(days=1)),
            item("Вчера", today - timedelta(days=1), today),
        ],
        [
            item("7 дней", today - timedelta(days=6), today + timedelta(days=1)),
            item("30 дней", today - timedelta(days=29), today + timedelta(days=1)),
        ],
        [
            item("Этот месяц", first, add_months(first, 1)),
            item("Прошлый месяц", prev_first, first),
        ],
        [
            item("Квартал", quarter, add_months(quarter, 3)),
            item("Год", date(today.year, 1, 1), date(today.year + 1, 1, 1)),
        ],
        [
            InlineKeyboardButton(
                "\U0001f4c5 Выбрать месяц", callback_data=f"period:months:{action}:{today.year}"
            ),
            InlineKeyboardButton(
                "✏️ Свой период", callback_data=f"period:custom:{action}"
            ),
        ],
    ]
    rows.extend(extra_rows or [])
    rows.append([InlineKeyboardButton("⬅️ Главное меню", callback_data="nav:main")])
    return InlineKeyboardMarkup(rows)


def kb_period_months(action: str, year: int) -> InlineKeyboardMarkup:
    """Сетка 12 месяцев выбранного года + переключение года."""
    rows, row = [], []
    for month in range(1, 13):
        start, end = month_period(year, month)
        row.append(
            InlineKeyboardButton(
                f"{month:02d}",
                callback_data=f"period:go:{action}:{start.isoformat()}:{end.isoformat()}",
            )
        )
        if len(row) == 4:
            rows.append(row)
            row = []
    if row:
        rows.append(row)
    rows.append(
        [
            InlineKeyboardButton(
                f"⬅️ {year - 1}", callback_data=f"period:months:{action}:{year - 1}"
            ),
            InlineKeyboardButton(f"{year}", callback_data="nav:noop"),
            InlineKeyboardButton(
                f"{year + 1} ➡️", callback_data=f"period:months:{action}:{year + 1}"
            ),
        ]
    )
    rows.append(
        [
            InlineKeyboardButton(
                f"Весь {year} год",
                callback_data=(
                    f"period:go:{action}:{date(year, 1, 1).isoformat()}:"
                    f"{date(year + 1, 1, 1).isoformat()}"
                ),
            )
        ]
    )
    rows.append(
        [InlineKeyboardButton("⬅️ Назад", callback_data=f"period:pick:{action}")]
    )
    return InlineKeyboardMarkup(rows)


# -----------------------------------------------------------------------------
# Проверяемые расчёты: считает Питон, а не модель
# -----------------------------------------------------------------------------
def calc_sheet(
    label: str,
    rows: list[dict],
    cats: list[dict] | None = None,
    prev_rows: list[dict] | None = None,
) -> tuple[list[str], set[Decimal]]:
    """Строки вида «Чистыми = 150 000,00 − 45 000,00 = 105 000,00».

    Их видит и модель (как «ПРОВЕРЕННЫЕ РАСЧЁТЫ»), и владелец в кнопке
    «\U0001f9fe Как посчитано», так что любую цифру можно перепроверить вручную.
    """
    zero = Decimal("0.00")
    revenue = sum((r["revenue"] for r in rows), zero)
    expenses = sum((r["expenses"] for r in rows), zero)
    net = revenue - expenses
    shares = sum((r["artist_share"] for r in rows), zero)
    label_profit = net - shares
    tax = (revenue * Config.TAX_RATE).quantize(CENTS, rounding=ROUND_HALF_UP)

    lines = [f"Период: {label}"]
    lines.append(f"Выручка = сумма всех доходов = {money(revenue)}")
    lines.append(f"Расходы = сумма всех расходов = {money(expenses)}")
    lines.append(f"Чистыми = {money(revenue)} − {money(expenses)} = {money(net)}")
    if revenue > 0:
        margin = (net / revenue * 100).quantize(Decimal("0.1"))
        lines.append(
            f"Маржа = {money(net)} ÷ {money(revenue)} × 100 = {margin}%"
        )
    for r in rows:
        rate = Decimal(str(r["rate"] or 0))
        lines.append(
            f"{r['artist']}: ({money(r['revenue'])} − {money(r['expenses'])}) "
            f"× {rate} = {money(r['artist_share'])} артисту; "
            f"лейблу {money(r['gross'])} − {money(r['artist_share'])} "
            f"= {money(r['label_profit'])}"
        )
    lines.append(
        "Итого артистам = "
        + (" + ".join(money(r["artist_share"]) for r in rows) or "0,00")
        + f" = {money(shares)}"
    )
    lines.append(
        f"Прибыль лейбла = {money(net)} − {money(shares)} = {money(label_profit)}"
    )
    lines.append(
        f"Налог {(Config.TAX_RATE * 100).normalize()}% = {money(revenue)} "
        f"× {Config.TAX_RATE} = {money(tax)}"
    )
    if prev_rows is not None:
        prev_revenue = sum((r["revenue"] for r in prev_rows), zero)
        delta = revenue - prev_revenue
        lines.append(
            f"Динамика выручки = {money(revenue)} − {money(prev_revenue)} "
            f"= {money(delta)}"
        )
    for c in (cats or [])[:10]:
        kind = "доход" if c["kind"] == "income" else "расход"
        base = revenue if c["kind"] == "income" else expenses
        share = (
            f" = {(c['total'] / base * 100).quantize(Decimal('0.1'))}% от {money(base)}"
            if base > 0
            else ""
        )
        lines.append(
            f"{kind} «{c['category']}»: {money(c['total'])} за {c['count']} опер.{share}"
        )

    numbers = collect_numbers(lines)
    numbers |= {revenue, expenses, net, shares, label_profit, tax}
    return lines, numbers


def calc_sheet_html(label: str, lines: list[str]) -> str:
    body = "\n".join(f"• <code>{line}</code>" for line in lines[1:])
    return (
        f"\U0001f9fe <b>Как посчитано · {label}</b>\n"
        "<i>Считал код бота, а не модель — этим цифрам можно верить.</i>\n\n"
        f"{body}"
    )




# -----------------------------------------------------------------------------
# Клавиатуры (интерфейс только на inline-кнопках)
# -----------------------------------------------------------------------------
# -----------------------------------------------------------------------------
# Роли и права
#   owner  — всё + управление доступом и ролями
#   admin  — всё по финансам + приглашение наблюдателей
#   viewer — только чтение: сводки, отчёты, аномалии, вопросы ИИ
# -----------------------------------------------------------------------------
ROLE_OWNER = "owner"
ROLE_ADMIN = "admin"
ROLE_VIEWER = "viewer"
ROLE_TITLES = {
    ROLE_OWNER: "👑 владелец",
    ROLE_ADMIN: "🛠 админ",
    ROLE_VIEWER: "👀 наблюдатель",
}
# Роль текущего апдейта: меню рисуется по правам без передачи аргументов.
CURRENT_ROLE: ContextVar[str] = ContextVar("current_role", default=ROLE_VIEWER)

# Действия, меняющие данные — недоступны наблюдателю.
WRITE_GATED = {
    ("artist", "add"),
    ("tx", "new"),
    ("tx", "artist"),
    ("tx", "undo"),
    ("tx", "drop"),
    ("excel", "wait"),
    ("pay", "paid"),
    ("quick", "hint"),
    ("quick", "save"),
}


def can_write(role: str | None = None) -> bool:
    return (role or CURRENT_ROLE.get()) in (ROLE_OWNER, ROLE_ADMIN)


def can_manage(role: str | None = None) -> bool:
    return (role or CURRENT_ROLE.get()) in (ROLE_OWNER, ROLE_ADMIN)


def user_title(u: dict) -> str:
    if u.get("full_name"):
        return str(u["full_name"])
    if u.get("username"):
        return "@" + str(u["username"])
    return f"id {u['tg_user_id']}"


def kb_main() -> InlineKeyboardMarkup:
    role = CURRENT_ROLE.get()
    write = can_write(role)
    rows: list[list[InlineKeyboardButton]] = [
        [InlineKeyboardButton("💬 Спросить бухгалтера", callback_data="ai:ask")]
    ]
    if write:
        rows.append(
            [
                InlineKeyboardButton("➕ Доход", callback_data="tx:new:income"),
                InlineKeyboardButton("➖ Расход", callback_data="tx:new:expense"),
            ]
        )
    rows.append(
        [
            InlineKeyboardButton("📈 Сводка", callback_data="dash:now"),
            InlineKeyboardButton("📊 Отчёт за месяц", callback_data="report:menu"),
        ]
    )
    rows.append(
        [
            InlineKeyboardButton("🚨 Аномалии", callback_data="anomaly:run"),
            InlineKeyboardButton("💸 Платежи", callback_data="pay:list:all"),
        ]
    )
    rows.append(
        [
            InlineKeyboardButton("🧾 Как посчитано", callback_data="check:calc"),
            InlineKeyboardButton("🔍 Найти операцию", callback_data="find:start"),
        ]
    )
    rows.append(
        [
            InlineKeyboardButton("⚖️ Сравнить периоды", callback_data="cmp:menu"),
            InlineKeyboardButton("🧠 Модель ИИ", callback_data="model:menu"),
        ]
    )
    rows.append(
        [
            InlineKeyboardButton("🩺 Проверка данных", callback_data="audit:run"),
            InlineKeyboardButton("📅 Сводка за период", callback_data="period:pick:dash"),
        ]
    )
    rows.append([InlineKeyboardButton("🎨 Генерация", callback_data="gen:menu")])
    if write:
        rows.append(
            [
                InlineKeyboardButton(
                    "📥 Импорт документа", callback_data="excel:wait"
                ),
                InlineKeyboardButton(
                    "📤 Экспорт документа", callback_data="export:menu"
                ),
            ]
        )
    else:
        rows.append(
            [
                InlineKeyboardButton(
                    "📤 Экспорт документа", callback_data="export:menu"
                )
            ]
        )
    artist_row = [InlineKeyboardButton("🎤 Артисты", callback_data="artist:list")]
    if write:
        artist_row.append(
            InlineKeyboardButton("👤 Новый артист", callback_data="artist:add")
        )
    rows.append(artist_row)
    if write:
        rows.append(
            [
                InlineKeyboardButton(
                    "↩️ Отменить последнюю операцию", callback_data="tx:undo"
                )
            ]
        )
    if can_manage(role):
        rows.append(
            [InlineKeyboardButton("🔑 Доступ и люди", callback_data="access:menu")]
        )
    return InlineKeyboardMarkup(rows)


def kb_access(role: str) -> InlineKeyboardMarkup:
    rows = [
        [InlineKeyboardButton("👥 Кто имеет доступ", callback_data="access:users")],
        [
            InlineKeyboardButton(
                "🎫 Код для наблюдателя", callback_data="access:invite:viewer"
            )
        ],
    ]
    if role == ROLE_OWNER:
        rows.append(
            [
                InlineKeyboardButton(
                    "🎫 Код для админа", callback_data="access:invite:admin"
                )
            ]
        )
    rows.append(
        [InlineKeyboardButton("🔗 Активные коды", callback_data="access:codes")]
    )
    rows.append(
        [
            InlineKeyboardButton(
                "➕ Выдать по Telegram ID", callback_data="access:grant"
            )
        ]
    )
    rows.append(
        [InlineKeyboardButton("⬅️ Главное меню", callback_data="nav:main")]
    )
    return InlineKeyboardMarkup(rows)


def kb_users(users: list[dict]) -> InlineKeyboardMarkup:
    rows = []
    for u in users[:40]:
        mark = "" if int(u["is_active"]) else "🚫 "
        short = ROLE_TITLES[u["role"]].split()[-1]
        rows.append(
            [
                InlineKeyboardButton(
                    f"{mark}{user_title(u)} — {short}",
                    callback_data=f"access:user:{u['tg_user_id']}",
                )
            ]
        )
    rows.append([InlineKeyboardButton("⬅️ Назад", callback_data="access:menu")])
    return InlineKeyboardMarkup(rows)


def manageable(row: dict, viewer_role: str, me: int) -> bool:
    if int(row["tg_user_id"]) == me:
        return False
    if row["role"] == ROLE_OWNER:
        return False
    return viewer_role == ROLE_OWNER or row["role"] == ROLE_VIEWER


def kb_user(row: dict, viewer_role: str, me: int) -> InlineKeyboardMarkup:
    uid = int(row["tg_user_id"])
    rows: list[list[InlineKeyboardButton]] = []
    if manageable(row, viewer_role, me):
        if viewer_role == ROLE_OWNER:
            if row["role"] == ROLE_ADMIN:
                rows.append(
                    [
                        InlineKeyboardButton(
                            "👀 Сделать наблюдателем",
                            callback_data=f"access:role:{uid}:viewer",
                        )
                    ]
                )
            else:
                rows.append(
                    [
                        InlineKeyboardButton(
                            "🛠 Сделать админом",
                            callback_data=f"access:role:{uid}:admin",
                        )
                    ]
                )
        if int(row["is_active"]):
            rows.append(
                [
                    InlineKeyboardButton(
                        "🚫 Забрать доступ", callback_data=f"access:block:{uid}"
                    )
                ]
            )
        else:
            rows.append(
                [
                    InlineKeyboardButton(
                        "✅ Вернуть доступ", callback_data=f"access:unblock:{uid}"
                    )
                ]
            )
        if viewer_role == ROLE_OWNER:
            rows.append(
                [
                    InlineKeyboardButton(
                        "🗑 Удалить из списка",
                        callback_data=f"access:delete:{uid}",
                    )
                ]
            )
    rows.append([InlineKeyboardButton("⬅️ К списку", callback_data="access:users")])
    return InlineKeyboardMarkup(rows)


def kb_codes(invites: list[dict]) -> InlineKeyboardMarkup:
    rows = []
    for inv in invites[:20]:
        rows.append(
            [
                InlineKeyboardButton(
                    f"❌ Отозвать {inv['code']}",
                    callback_data=f"access:kill:{inv['code']}",
                )
            ]
        )
    rows.append([InlineKeyboardButton("⬅️ Назад", callback_data="access:menu")])
    return InlineKeyboardMarkup(rows)


def kb_back(
    extra: list[list[InlineKeyboardButton]] | None = None,
) -> InlineKeyboardMarkup:
    rows = list(extra or [])
    rows.append([InlineKeyboardButton("⬅️ Главное меню", callback_data="nav:main")])
    return InlineKeyboardMarkup(rows)


def kb_artists(
    artists: list[dict], prefix: str, allow_none: bool = False
) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    for i in range(0, len(artists), 2):
        rows.append(
            [
                InlineKeyboardButton(
                    str(a["name"])[:24], callback_data=f"{prefix}:{a['id']}"
                )
                for a in artists[i : i + 2]
            ]
        )
    if allow_none:
        rows.append([InlineKeyboardButton(NO_ARTIST, callback_data=f"{prefix}:0")])
    rows.append([InlineKeyboardButton("⬅️ Главное меню", callback_data="nav:main")])
    return InlineKeyboardMarkup(rows)


def kb_months() -> InlineKeyboardMarkup:
    today = date.today()
    rows, row = [], []
    for offset in range(6):
        month = today.month - offset
        year = today.year
        while month <= 0:
            month += 12
            year -= 1
        row.append(
            InlineKeyboardButton(
                f"{month:02d}.{year}", callback_data=f"report:run:{year}:{month}"
            )
        )
        if len(row) == 3:
            rows.append(row)
            row = []
    if row:
        rows.append(row)
    rows.append([InlineKeyboardButton("⬅️ Главное меню", callback_data="nav:main")])
    return InlineKeyboardMarkup(rows)


def kb_payments(payments: list[dict]) -> InlineKeyboardMarkup:
    rows = [
        [
            InlineKeyboardButton(
                f"✅ Оплатить #{p['id']} · {str(p['artist'])[:14]}",
                callback_data=f"pay:paid:{p['id']}",
            )
        ]
        for p in payments
        if p["status"] == "pending"
    ][:8]
    rows.append(
        [
            InlineKeyboardButton("⏳ Ожидают", callback_data="pay:list:pending"),
            InlineKeyboardButton("✅ Оплачены", callback_data="pay:list:paid"),
        ]
    )
    rows.append(
        [InlineKeyboardButton("🔄 Пересчитать за месяц", callback_data="pay:recalc")]
    )
    rows.append([InlineKeyboardButton("⬅️ Главное меню", callback_data="nav:main")])
    return InlineKeyboardMarkup(rows)


# -----------------------------------------------------------------------------
# Доступ и вспомогательные ответы
# -----------------------------------------------------------------------------
def _role_from_env(user_id: int) -> str | None:
    """Поддержка старых переменных: OWNER_IDS и ALLOWED_USER_IDS."""
    if user_id in Config.OWNER_IDS:
        return ROLE_OWNER
    if user_id in Config.ALLOWED_USER_IDS:
        return ROLE_ADMIN
    return None


async def resolve_access(update: Update, repo: "Repo") -> dict | None:
    """Кто пишет боту и с какой ролью. None — доступа нет."""
    user = update.effective_user
    if not user:
        return None

    row = await asyncio.to_thread(
        repo.touch_user, user.id, user.username, user.full_name
    )

    if row is None:
        role = _role_from_env(user.id)
        if role is None and not Config.OWNER_IDS and not Config.ALLOWED_USER_IDS:
            # Первый запуск без OWNER_IDS: первый человек становится владельцем.
            if await asyncio.to_thread(repo.users_count) == 0:
                role = ROLE_OWNER
                log.warning(
                    "Первый пользователь %s назначен владельцем бота", user.id
                )
        if role is None:
            return None
        row = await asyncio.to_thread(
            repo.grant, user.id, role, None, user.username, user.full_name
        )

    if row is None or not int(row["is_active"]):
        return None

    CURRENT_ROLE.set(row["role"])
    return row


async def handle_locked_message(
    update: Update, repo: "Repo", text: str
) -> None:
    """Сообщение от человека без доступа: пробуем текст как код-приглашение."""
    user = update.effective_user
    candidate = text.strip()
    if candidate.lower().startswith("/start"):  # диплинк t.me/<bot>?start=CODE
        candidate = candidate[6:].strip()
    candidate = candidate.replace("-", "").replace(" ", "").upper()

    if 6 <= len(candidate) <= 12 and candidate.isalnum():
        row = await asyncio.to_thread(
            repo.redeem_invite, candidate, user.id, user.username, user.full_name
        )
        if row:
            CURRENT_ROLE.set(row["role"])
            await show_main_menu(
                update,
                f"✅ Доступ открыт. Ваша роль: {ROLE_TITLES[row['role']]}.",
            )
            return
        await update.effective_chat.send_message(
            "⚠️ Код не подходит: неверный, просроченный или уже использованный."
        )
        return

    await update.effective_chat.send_message(
        "🔒 <b>Закрытый бот</b>\n\n"
        "Попросите у владельца код-приглашение и пришлите его сюда одним "
        "сообщением, например <code>K7P2M9QA</code>.\n\n"
        f"Ваш Telegram ID: <code>{user.id}</code>",
        parse_mode=ParseMode.HTML,
    )


async def send(
    update: Update, text: str, keyboard: InlineKeyboardMarkup | None = None
) -> None:
    parts = chunks(text)
    chat = update.effective_chat
    for idx, part in enumerate(parts):
        markup = keyboard if idx == len(parts) - 1 else None
        try:
            await chat.send_message(part, reply_markup=markup, parse_mode=ParseMode.HTML)
        except BadRequest:  # некорректный HTML от модели — отправляем как текст
            await chat.send_message(part, reply_markup=markup)


async def show_main_menu(update: Update, prefix: str = "") -> None:
    head = f"{prefix}\n\n" if prefix else ""
    # подвал с меткой сборки: сразу видно, какая версия реально задеплоена
    model = ACTIVE_MODEL.get("") or MODEL_CATALOG[base_model()]["title"]
    footer = f"\n\n<code>build {BUILD} · {model}</code>"
    await send(
        update,
        f"{head}<b>💼 Финансы лейбла {Config.LABEL_NAME}</b>\nВыберите действие:{footer}",
        kb_main(),
    )


def reset_state(context: ContextTypes.DEFAULT_TYPE) -> None:
    for key in ("state", "draft"):
        context.user_data.pop(key, None)


def prev_month(year: int, month: int) -> tuple[int, int]:
    return (year - 1, 12) if month == 1 else (year, month - 1)


def kb_export_months() -> InlineKeyboardMarkup:
    today = date.today()
    rows, row = [], []
    for offset in range(6):
        month, year = today.month - offset, today.year
        while month <= 0:
            month += 12
            year -= 1
        row.append(
            InlineKeyboardButton(
                f"{month:02d}.{year}", callback_data=f"export:month:{year}:{month}"
            )
        )
        if len(row) == 3:
            rows.append(row)
            row = []
    if row:
        rows.append(row)
    rows.append(
        [InlineKeyboardButton("⬅️ Главное меню", callback_data="nav:main")]
    )
    return InlineKeyboardMarkup(rows)


def pct_delta(now: Decimal, before: Decimal) -> str:
    if before == 0:
        return "новое" if now else "—"
    change = (now - before) / before * 100
    arrow = "📈" if change >= 0 else "📉"
    return f"{arrow} {change:+.0f}%"


def dashboard_text(
    label: str,
    prev_label: str,
    days: int,
    cur: dict,
    before: dict,
    breakdown: list[dict],
    cats: list[dict],
) -> str:
    """Сводка за ЛЮБОЙ период с расшифровкой каждой цифры."""
    cur_currency = Config.DEFAULT_CURRENCY
    to_artists = sum(
        (Decimal(str(r.get("artist_share") or 0)) for r in breakdown), Decimal("0")
    )
    label_profit = sum(
        (Decimal(str(r.get("label_profit") or 0)) for r in breakdown), Decimal("0")
    )
    margin = (cur["net"] / cur["revenue"] * 100) if cur["revenue"] else Decimal("0")
    tax = (cur["revenue"] * Config.TAX_RATE).quantize(Decimal("0.01"))
    days = max(int(days), 1)
    burn = (cur["expenses"] / days).quantize(Decimal("0.01"))

    top_expense = sorted(
        [c for c in cats if str(c.get("kind")) == "expense"],
        key=lambda c: Decimal(str(c.get("total") or 0)),
        reverse=True,
    )[:3]
    top_artists = sorted(
        breakdown,
        key=lambda r: Decimal(str(r.get("label_profit") or 0)),
        reverse=True,
    )[:3]

    lines = [
        f"\U0001f4c8 <b>Сводка · {label}</b>",
        "",
        f"Выручка: <b>{money(cur['revenue'], cur_currency)}</b> "
        f"({pct_delta(cur['revenue'], before['revenue'])} к {prev_label})",
        f"Расходы: <b>{money(cur['expenses'], cur_currency)}</b> "
        f"({pct_delta(cur['expenses'], before['expenses'])})",
        f"Чистыми: <b>{money(cur['net'], cur_currency)}</b> "
        f"<code>{money(cur['revenue'])} − {money(cur['expenses'])} "
        f"= {money(cur['net'])}</code>",
        f"Маржа: <b>{margin:.0f}%</b> "
        f"<code>{money(cur['net'])} ÷ {money(cur['revenue'])} × 100</code>",
        "",
        f"К выплате артистам: <b>{money(to_artists, cur_currency)}</b>",
        f"Остаётся лейблу: <b>{money(label_profit, cur_currency)}</b> "
        f"<code>{money(cur['net'])} − {money(to_artists)} = {money(label_profit)}</code>",
        f"Налоговый резерв ({Config.TAX_RATE * 100:.0f}%): "
        f"<b>{money(tax, cur_currency)}</b> "
        f"<code>{money(cur['revenue'])} × {Config.TAX_RATE}</code>",
        f"Средний расход в день: {money(burn, cur_currency)} "
        f"<code>{money(cur['expenses'])} ÷ {days}</code>",
        f"Операций за период: {cur['count']}",
    ]
    if top_expense:
        lines += ["", "<b>Куда уходят деньги:</b>"]
        lines += [
            f"• {c.get('category') or 'без категории'} — "
            f"{money(c.get('total'), cur_currency)}"
            for c in top_expense
        ]
    if top_artists:
        lines += ["", "<b>Лучшие по прибыли лейбла:</b>"]
        lines += [
            f"• {r.get('artist')}: ({money(r.get('revenue'))} − "
            f"{money(r.get('expenses'))}) × {Decimal(str(r.get('rate') or 0))} = "
            f"{money(r.get('artist_share'))} артисту, лейблу "
            f"{money(r.get('label_profit'))}"
            for r in top_artists
        ]
    lines += ["", "<i>Все цифры посчитаны кодом бота, не моделью.</i>"]
    return "\n".join(lines)


# -----------------------------------------------------------------------------
# Генерация документов и фотоотчёта — без ИИ, 0 токенов
# -----------------------------------------------------------------------------
DOC_FORMATS: dict[str, dict[str, str]] = {
    "xlsx": {
        "title": "Excel (.xlsx)",
        "icon": "\U0001f4d7",
        "about": "таблица с живыми формулами: сводка, артисты, категории, операции",
    },
    "docx": {
        "title": "Word (.docx)",
        "icon": "\U0001f4d8",
        "about": "готовый отчёт текстом — можно сразу отправить партнёру",
    },
    "csv": {
        "title": "CSV (.csv)",
        "icon": "\U0001f4c4",
        "about": "сырые операции для бухгалтерии или 1С",
    },
}


def report_filename(prefix: str, start: date, end: date, ext: str) -> str:
    last = end - timedelta(days=1)
    return f"{prefix}-{start.isoformat()}-{last.isoformat()}.{ext}"


def report_totals(rows: list[dict]) -> dict[str, Decimal]:
    """Итоги по разбору артистов — те же формулы, что и в «Как посчитано»."""
    zero = Decimal("0.00")
    revenue = sum((r["revenue"] for r in rows), zero)
    expenses = sum((r["expenses"] for r in rows), zero)
    shares = sum((r["artist_share"] for r in rows), zero)
    net = revenue - expenses
    tax = (revenue * Config.TAX_RATE).quantize(CENTS, rounding=ROUND_HALF_UP)
    margin = (
        (net / revenue * 100).quantize(Decimal("0.1")) if revenue > 0 else Decimal("0.0")
    )
    return {
        "revenue": revenue,
        "expenses": expenses,
        "net": net,
        "shares": shares,
        "label_profit": net - shares,
        "tax": tax,
        "margin": margin,
    }


def xlsx_report(
    label: str,
    breakdown: list[dict],
    cats: list[dict],
    tx_rows: list[dict],
    calc_lines: list[str],
) -> bytes:
    """Excel с живыми формулами: любую цифру можно пересчитать в самом файле."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    money_fmt = "#,##0.00"
    head_fill = PatternFill("solid", fgColor="1F2933")
    head_font = Font(bold=True, color="FFFFFF")

    wb = Workbook()

    def style_head(ws: Any, width_map: dict[str, int]) -> None:
        for cell in ws[1]:
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column, width in width_map.items():
            ws.column_dimensions[column].width = width
        ws.freeze_panes = "A2"

    summary = wb.active
    summary.title = "Сводка"

    artists = wb.create_sheet("Артисты")
    artists.append(
        [
            "Артист",
            "Выручка",
            "Расходы",
            "Ставка",
            "Чистыми",
            "Артисту",
            "Лейблу",
            "Операций",
        ]
    )
    for idx, row in enumerate(breakdown, start=2):
        artists.append(
            [
                row["artist"],
                float(row["revenue"]),
                float(row["expenses"]),
                float(row["rate"] or 0),
                f"=B{idx}-C{idx}",
                f"=(B{idx}-C{idx})*D{idx}",
                f"=E{idx}-F{idx}",
                int(row["tx_count"]),
            ]
        )
        for column in ("B", "C", "E", "F", "G"):
            artists[f"{column}{idx}"].number_format = money_fmt
        artists[f"D{idx}"].number_format = "0.00%" if row["rate"] and row["rate"] < 1 else "0.00"
    last_artist = max(2, len(breakdown) + 1)
    style_head(
        artists,
        {"A": 26, "B": 15, "C": 15, "D": 10, "E": 15, "F": 15, "G": 15, "H": 11},
    )

    categories = wb.create_sheet("Категории")
    categories.append(["Тип", "Категория", "Сумма", "Операций", "Доля в типе"])
    for idx, cat in enumerate(cats, start=2):
        kind = "доход" if cat["kind"] == "income" else "расход"
        categories.append(
            [
                kind,
                cat["category"],
                float(cat["total"]),
                int(cat["count"]),
                f'=IFERROR(C{idx}/SUMIF($A$2:$A${max(2, len(cats) + 1)},A{idx},'
                f'$C$2:$C${max(2, len(cats) + 1)}),0)',
            ]
        )
        categories[f"C{idx}"].number_format = money_fmt
        categories[f"E{idx}"].number_format = "0.0%"
    style_head(categories, {"A": 10, "B": 28, "C": 16, "D": 12, "E": 14})

    operations = wb.create_sheet("Операции")
    operations.append(
        [
            "id",
            "Дата",
            "Тип",
            "Сумма",
            "Валюта",
            "Артист",
            "Категория",
            "Описание",
            "Источник",
        ]
    )
    for idx, row in enumerate(tx_rows, start=2):
        operations.append(
            [
                row["id"],
                str(row["occurred_on"])[:10],
                "доход" if row["kind"] == "income" else "расход",
                float(row["amount"]),
                row.get("currency") or Config.DEFAULT_CURRENCY,
                row.get("artist") or "",
                row.get("category") or "",
                (row.get("description") or "")[:200],
                row.get("source") or "",
            ]
        )
        operations[f"D{idx}"].number_format = money_fmt
    style_head(
        operations,
        {"A": 7, "B": 12, "C": 9, "D": 15, "E": 9, "F": 22, "G": 20, "H": 40, "I": 12},
    )

    proof = wb.create_sheet("Как посчитано")
    proof.append(["Расчёт (посчитан кодом бота, не моделью)"])
    for line in calc_lines:
        proof.append([line])
    style_head(proof, {"A": 110})

    summary["A1"] = f"Отчёт лейбла {Config.LABEL_NAME}"
    summary["A1"].font = Font(bold=True, size=16)
    summary["A2"] = f"Период: {label}"
    summary["A3"] = f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')} · build {BUILD}"
    summary["A4"] = "Все цифры — живые формулы Excel, можно проверить и пересчитать."

    metrics = [
        ("Выручка", f"=SUM(Артисты!B2:B{last_artist})", "сумма всех доходов"),
        ("Расходы", f"=SUM(Артисты!C2:C{last_artist})", "сумма всех расходов"),
        ("Чистыми", "=B6-B7", "Выручка − Расходы"),
        ("Артистам", f"=SUM(Артисты!F2:F{last_artist})", "(Выручка − Расходы) × Ставка"),
        ("Прибыль лейбла", "=B8-B9", "Чистыми − выплаты артистам"),
        (
            f"Налоговый резерв {(Config.TAX_RATE * 100).normalize()}%",
            f"=B6*{float(Config.TAX_RATE)}",
            "Выручка × ставка налога",
        ),
        ("Маржа", "=IFERROR(B8/B6,0)", "Чистыми ÷ Выручка"),
        ("Операций", f"=SUM(Артисты!H2:H{last_artist})", "сколько записей в периоде"),
    ]
    summary["A5"] = "Показатель"
    summary["B5"] = "Значение"
    summary["C5"] = "Как посчитано"
    for cell in ("A5", "B5", "C5"):
        summary[cell].fill = head_fill
        summary[cell].font = head_font
    for offset, (name, formula, how) in enumerate(metrics):
        line = 6 + offset
        summary[f"A{line}"] = name
        summary[f"B{line}"] = formula
        summary[f"C{line}"] = how
        summary[f"B{line}"].number_format = (
            "0.0%" if name == "Маржа" else ("0" if name == "Операций" else money_fmt)
        )
    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 18
    summary.column_dimensions["C"].width = 42

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _xml_escape(text: Any) -> str:
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _docx_paragraph(text: str, size: int = 22, bold: bool = False) -> str:
    props = f'<w:rPr>{"<w:b/>" if bold else ""}<w:sz w:val="{size}"/>'
    props += f'<w:szCs w:val="{size}"/></w:rPr>'
    return (
        "<w:p><w:pPr><w:spacing w:after=\"120\"/></w:pPr>"
        f"<w:r>{props}<w:t xml:space=\"preserve\">{_xml_escape(text)}</w:t></w:r></w:p>"
    )


def docx_report(title: str, blocks: list[tuple[str, str]]) -> bytes:
    """Минимальный валидный .docx без внешних библиотек (Word/Google Docs открывают)."""
    import zipfile

    body = [_docx_paragraph(title, size=34, bold=True)]
    for kind, text in blocks:
        if kind == "h":
            body.append(_docx_paragraph(text, size=28, bold=True))
        elif kind == "li":
            body.append(_docx_paragraph(f"\u2022 {text}", size=22))
        elif kind == "small":
            body.append(_docx_paragraph(text, size=18))
        else:
            body.append(_docx_paragraph(text, size=22))

    document = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        "<w:body>" + "".join(body) + '<w:sectPr><w:pgSz w:w="11906" w:h="16838"/>'
        '<w:pgMar w:top="1134" w:right="850" w:bottom="1134" w:left="1134"/>'
        "</w:sectPr></w:body></w:document>"
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" '
        'ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" ContentType="application/vnd.'
        'openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        "</Types>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/'
        '2006/relationships/officeDocument" Target="word/document.xml"/>'
        "</Relationships>"
    )

    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", rels)
        zf.writestr("word/document.xml", document)
    return buf.getvalue()


def report_blocks(
    label: str,
    breakdown: list[dict],
    cats: list[dict],
    calc_lines: list[str],
    tx_count: int,
) -> list[tuple[str, str]]:
    """Текст отчёта для Word: только посчитанные кодом цифры."""
    totals = report_totals(breakdown)
    cur = Config.DEFAULT_CURRENCY
    blocks: list[tuple[str, str]] = [
        ("small", f"Период: {label} · операций: {tx_count}"),
        (
            "small",
            f"Сформирован ботом {datetime.now().strftime('%d.%m.%Y %H:%M')} · build {BUILD}",
        ),
        ("h", "1. Итоги"),
        ("li", f"Выручка: {money(totals['revenue'], cur)}"),
        ("li", f"Расходы: {money(totals['expenses'], cur)}"),
        (
            "li",
            f"Чистыми: {money(totals['revenue'])} − {money(totals['expenses'])} "
            f"= {money(totals['net'], cur)}",
        ),
        ("li", f"Маржа: {totals['margin']}%"),
        ("li", f"К выплате артистам: {money(totals['shares'], cur)}"),
        (
            "li",
            f"Прибыль лейбла: {money(totals['net'])} − {money(totals['shares'])} "
            f"= {money(totals['label_profit'], cur)}",
        ),
        (
            "li",
            f"Налоговый резерв {(Config.TAX_RATE * 100).normalize()}%: "
            f"{money(totals['tax'], cur)}",
        ),
        ("h", "2. Артисты"),
    ]
    if breakdown:
        for row in breakdown[:20]:
            blocks.append(
                (
                    "li",
                    f"{row['artist']}: ({money(row['revenue'])} − {money(row['expenses'])})"
                    f" × {row['rate']} = {money(row['artist_share'])} артисту, "
                    f"лейблу {money(row['label_profit'])}",
                )
            )
    else:
        blocks.append(("p", "За период операций нет."))

    blocks.append(("h", "3. Категории"))
    if cats:
        for cat in cats[:15]:
            kind = "доход" if cat["kind"] == "income" else "расход"
            blocks.append(
                (
                    "li",
                    f"{kind} «{cat['category']}»: {money(cat['total'], cur)} "
                    f"за {cat['count']} опер.",
                )
            )
    else:
        blocks.append(("p", "Категорий нет."))

    blocks.append(("h", "4. Как посчитано (проверьте любую цифру)"))
    for line in calc_lines:
        blocks.append(("li", line))
    blocks.append(
        (
            "small",
            "Все суммы посчитаны кодом бота из базы данных, без участия ИИ.",
        )
    )
    return blocks


# ---------------------------------------------------------------------------
# Фотоотчёт: рисуем PNG сами, без генерации картинок через ИИ
# ---------------------------------------------------------------------------
FONT_CANDIDATES = (
    os.getenv("FONT_PATH", "").strip(),
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    "/usr/share/fonts/TTF/DejaVuSans.ttf",
    "/usr/share/fonts/dejavu/DejaVuSans.ttf",
    "/Library/Fonts/Arial Unicode.ttf",
    "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
)

_FONT_CACHE: dict[int, Any] = {}
_FONT_PATH: list[str | None] = []

_TRANSLIT = {
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e", "ж": "zh",
    "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m", "н": "n", "о": "o",
    "п": "p", "р": "r", "с": "s", "т": "t", "у": "u", "ф": "f", "х": "h", "ц": "c",
    "ч": "ch", "ш": "sh", "щ": "sch", "ъ": "", "ы": "y", "ь": "", "э": "e",
    "ю": "yu", "я": "ya",
}


def _resolve_font_path() -> str | None:
    if _FONT_PATH:
        return _FONT_PATH[0]
    found: str | None = None
    for path in FONT_CANDIDATES:
        if path and os.path.exists(path):
            found = path
            break
    if found is None:
        # Шрифт из matplotlib, если он вдруг есть в образе (в нём полная кириллица).
        try:
            import matplotlib

            candidate = (
                Path(matplotlib.__file__).parent
                / "mpl-data"
                / "fonts"
                / "ttf"
                / "DejaVuSans.ttf"
            )
            if candidate.exists():
                found = str(candidate)
        except Exception:
            found = None
    _FONT_PATH.append(found)
    return found


def photo_text(text: str) -> str:
    """Если кириллического шрифта в системе нет — пишем транслитом, чтобы не было квадратов."""
    if _resolve_font_path():
        return text
    out = []
    for char in text:
        lower = char.lower()
        if lower in _TRANSLIT:
            repl = _TRANSLIT[lower]
            out.append(repl.upper() if char.isupper() else repl)
        else:
            out.append(char)
    return "".join(out)


def _font(size: int) -> Any:
    from PIL import ImageFont

    if size in _FONT_CACHE:
        return _FONT_CACHE[size]
    path = _resolve_font_path()
    font: Any
    if path:
        font = ImageFont.truetype(path, size)
    else:
        try:
            font = ImageFont.load_default(size=size)
        except TypeError:  # старые версии Pillow
            font = ImageFont.load_default()
    _FONT_CACHE[size] = font
    return font


def photo_report(
    label: str,
    breakdown: list[dict],
    cats: list[dict],
    tx_count: int,
    prev_totals: dict | None = None,
) -> bytes:
    """Картинка-отчёт с цифрами, графиком и формулами. Стоит 0 токенов."""
    try:
        from PIL import Image, ImageDraw
    except Exception as exc:  # pragma: no cover
        raise RuntimeError(
            "Для фотоотчёта нужен пакет Pillow — добавьте строку Pillow в "
            f"requirements.txt и нажмите «Пересобрать» ({exc})"
        ) from exc

    totals = report_totals(breakdown)
    cur = Config.DEFAULT_CURRENCY
    width, height = 1080, 1420
    bg = (17, 21, 28)
    card = (28, 34, 44)
    accent = (94, 214, 160)
    accent_bad = (232, 116, 116)
    text_main = (240, 244, 248)
    text_dim = (150, 160, 175)

    image = Image.new("RGB", (width, height), bg)
    draw = ImageDraw.Draw(image)

    def write(xy: tuple[int, int], text: str, size: int, color: tuple[int, int, int]) -> None:
        draw.text(xy, photo_text(text), font=_font(size), fill=color)

    draw.rectangle([0, 0, width, 190], fill=(23, 29, 38))
    write((48, 46), f"Лейбл {Config.LABEL_NAME} · фотоотчёт", 44, text_main)
    write((48, 106), f"Период: {label}", 32, accent)
    write((48, 148), f"Операций: {tx_count} · сформировано без ИИ (0 токенов)", 24, text_dim)

    kpis = [
        ("Выручка", money(totals["revenue"], cur), accent),
        ("Расходы", money(totals["expenses"], cur), accent_bad),
        (
            "Чистыми",
            money(totals["net"], cur),
            accent if totals["net"] >= 0 else accent_bad,
        ),
        ("Прибыль лейбла", money(totals["label_profit"], cur), text_main),
    ]
    top = 224
    for idx, (name, value, color) in enumerate(kpis):
        col, row = idx % 2, idx // 2
        x0 = 40 + col * 508
        y0 = top + row * 150
        draw.rounded_rectangle([x0, y0, x0 + 476, y0 + 130], radius=22, fill=card)
        write((x0 + 26, y0 + 22), name, 26, text_dim)
        write((x0 + 26, y0 + 62), value, 40, color)

    y = top + 2 * 150 + 24
    write((48, y), "Самые крупные статьи", 30, text_main)
    y += 48
    bars = sorted(cats, key=lambda c: -c["total"])[:6]
    top_value = max((c["total"] for c in bars), default=Decimal("0"))
    for cat in bars:
        share = float(cat["total"] / top_value) if top_value > 0 else 0.0
        color = accent if cat["kind"] == "income" else accent_bad
        draw.rounded_rectangle([48, y, 48 + int(760 * share) + 6, y + 44], radius=10, fill=color)
        kind = "доход" if cat["kind"] == "income" else "расход"
        write((62, y + 8), f"{cat['category']} ({kind})", 24, (18, 22, 30))
        write((830, y + 8), money(cat["total"]), 24, text_main)
        y += 58
    if not bars:
        write((48, y), "Нет данных за период", 26, text_dim)
        y += 58

    y += 14
    write((48, y), "Артисты: доля и прибыль лейбла", 30, text_main)
    y += 46
    for row in breakdown[:5]:
        write(
            (48, y),
            f"{row['artist']}: артисту {money(row['artist_share'])} · "
            f"лейблу {money(row['label_profit'])}",
            24,
            text_main,
        )
        y += 36
        write(
            (66, y),
            f"({money(row['revenue'])} - {money(row['expenses'])}) x {row['rate']}",
            22,
            text_dim,
        )
        y += 38

    y = max(y + 10, height - 250)
    draw.rounded_rectangle([40, y, width - 40, height - 40], radius=22, fill=card)
    write((66, y + 22), "Как посчитано", 28, accent)
    proof = [
        f"Чистыми = {money(totals['revenue'])} - {money(totals['expenses'])} "
        f"= {money(totals['net'])}",
        f"Артистам = {money(totals['shares'])}; лейблу = {money(totals['net'])} - "
        f"{money(totals['shares'])} = {money(totals['label_profit'])}",
        f"Маржа = {money(totals['net'])} / {money(totals['revenue'])} x 100 = "
        f"{totals['margin']}%",
        f"Налоговый резерв = {money(totals['revenue'])} x {Config.TAX_RATE} = "
        f"{money(totals['tax'])}",
    ]
    if prev_totals:
        delta = totals["revenue"] - Decimal(str(prev_totals.get("revenue") or 0))
        proof.append(
            f"Динамика выручки = {money(totals['revenue'])} - "
            f"{money(prev_totals.get('revenue') or 0)} = {money(delta)}"
        )
    line_y = y + 66
    for line in proof[:5]:
        write((66, line_y), line, 21, text_main)
        line_y += 32

    buf = BytesIO()
    image.save(buf, format="PNG", optimize=True)
    return buf.getvalue()


def transactions_csv(rows: list[dict]) -> bytes:
    buf = StringIO()
    writer = csv.writer(buf, delimiter=";")
    writer.writerow(
        ["id", "дата", "тип", "сумма", "валюта", "артист", "категория", "описание", "источник"]
    )
    for r in rows:
        writer.writerow(
            [
                r["id"],
                str(r["occurred_on"])[:10],
                "доход" if r["kind"] == "income" else "расход",
                f"{r['amount']:.2f}".replace(".", ","),
                r.get("currency") or Config.DEFAULT_CURRENCY,
                r.get("artist") or "",
                r.get("category") or "",
                r.get("description") or "",
                r.get("source") or "",
            ]
        )
    return buf.getvalue().encode("utf-8-sig")


async def collect_books(
    repo: "Repo",
    start: date | None = None,
    end: date | None = None,
    months: int = 3,
) -> dict:
    """Компактный срез базы за любой период для вопросов ИИ-бухгалтеру."""
    today = date.today()
    if start is None or end is None:
        start, end = month_period(today.year, today.month)
    span = max((end - start).days, 1)
    prev_start, prev_end = start - timedelta(days=span), start
    last, prev_last = end - timedelta(days=1), prev_end - timedelta(days=1)

    (
        cur,
        before,
        breakdown,
        prev_breakdown,
        cats,
        artists,
        recent,
        anomalies,
    ) = await asyncio.gather(
        asyncio.to_thread(repo.totals, start, last),
        asyncio.to_thread(repo.totals, prev_start, prev_last),
        asyncio.to_thread(repo.breakdown, start, end),
        asyncio.to_thread(repo.breakdown, prev_start, prev_end),
        asyncio.to_thread(repo.categories, start, end),
        asyncio.to_thread(repo.list_artists),
        asyncio.to_thread(repo.recent_transactions, 25),
        asyncio.to_thread(repo.anomaly_candidates, months),
    )
    label, prev_label = period_label(start, end), period_label(prev_start, prev_end)
    return {
        "сегодня": today.isoformat(),
        "валюта": Config.DEFAULT_CURRENCY,
        "период": label,
        "период_с": start.isoformat(),
        "период_по": last.isoformat(),
        "итоги_периода": {"период": label, **cur},
        "итоги_предыдущего_периода": {"период": prev_label, **before},
        "по_артистам": breakdown,
        "по_артистам_раньше": prev_breakdown,
        "по_категориям": cats,
        "артисты": [
            {"имя": a["name"], "ставка": str(a["rate"])} for a in artists
        ],
        "последние_операции": [
            {
                "дата": str(r["occurred_on"])[:10],
                "тип": r["kind"],
                "сумма": str(r["amount"]),
                "артист": r.get("artist"),
                "категория": r.get("category"),
                "описание": r.get("description"),
            }
            for r in recent
        ],
        "аномалии_и_целостность": anomalies,
        "налоговая_ставка": str(Config.TAX_RATE),
    }


def books_calc(books: dict) -> list[str]:
    """Проверяемые расчёты для среза данных."""
    lines, _ = calc_sheet(
        str(books.get("период") or ""),
        list(books.get("по_артистам") or []),
        list(books.get("по_категориям") or []),
        books.get("по_артистам_раньше"),
    )
    return lines


def remember_period(context: ContextTypes.DEFAULT_TYPE, start: date, end: date) -> None:
    context.user_data["period"] = [start.isoformat(), end.isoformat()]


def current_period(context: ContextTypes.DEFAULT_TYPE) -> tuple[date, date]:
    """Последний выбранный период; по умолчанию — текущий месяц."""
    raw = context.user_data.get("period")
    if isinstance(raw, list) and len(raw) == 2:
        try:
            return date.fromisoformat(raw[0]), date.fromisoformat(raw[1])
        except ValueError:
            pass
    today = date.today()
    return month_period(today.year, today.month)


def auto_mode(context: ContextTypes.DEFAULT_TYPE) -> bool:
    return bool(context.user_data.get("model_auto", Config.AI_MODE != "manual"))


def chosen_model(context: ContextTypes.DEFAULT_TYPE) -> str | None:
    """None => авторежим: бот сам выберет и при нужде поднимет модель."""
    if auto_mode(context):
        return None
    return normalize_model(context.user_data.get("model") or Config.AI_MANUAL_MODEL)


HARD_MARKERS = (
    "не сход",
    "расхожден",
    "аудит",
    "проверь всё",
    "проверь все",
    "пересчитай",
    "где ошибк",
    "налог",
    "за год",
    "квартал",
    "стратег",
)


def is_hard(text: str) -> bool:
    """Сложные запросы сразу получают доступ к тяжёлым моделям."""
    low = (text or "").lower()
    return any(marker in low for marker in HARD_MARKERS)


def kb_models(context: ContextTypes.DEFAULT_TYPE) -> InlineKeyboardMarkup:
    auto = auto_mode(context)
    manual = normalize_model(context.user_data.get("model") or Config.AI_MANUAL_MODEL)
    rows = [
        [
            InlineKeyboardButton(
                ("✅ " if auto else "") + "\U0001f916 Автовыбор (рекомендуется)",
                callback_data="model:auto",
            )
        ]
    ]
    for key, spec in MODEL_CATALOG.items():
        mark = "✅ " if (not auto and manual == key) else ""
        rows.append(
            [
                InlineKeyboardButton(
                    f"{mark}{spec['title']} {spec['price']}",
                    callback_data=f"model:set:{key}",
                ),
                InlineKeyboardButton("ℹ️ О модели", callback_data=f"model:info:{key}"),
            ]
        )
    rows.append([InlineKeyboardButton("⬅️ Главное меню", callback_data="nav:main")])
    return InlineKeyboardMarkup(rows)


AUTO_EXPLAIN = (
    "<b>Как работает автовыбор</b>\n"
    f"1. Обычные задачи считает <b>{MODEL_CATALOG['qwen']['title']}</b> — быстро и дешёво.\n"
    "2. Если он не уверен или появились цифры не из базы — подключаются "
    "дешёвые помощники <b>Sonnet 5</b> и <b>GPT 5.5</b>.\n"
    "3. Сложные разборы уходят к <b>Opus 5</b> или <b>GPT 5.6 Luna</b>.\n"
    "4. <b>Terra</b> и <b>Sol</b> включаются только в крайних случаях.\n\n"
    "Любую модель можно зафиксировать вручную кнопкой ниже."
)

PERIOD_HINT = (
    "\U0001f4c5 <b>За какой период?</b>\n"
    "Готовые варианты ниже, или выберите любой месяц и год, "
    "либо задайте свои даты."
)

CUSTOM_PERIOD_HINT = (
    "✏️ <b>Свой период</b>\nНапишите одним сообщением, как удобно:\n\n"
    "• <code>15.08.2026</code> — один день\n"
    "• <code>08.2026</code> или <code>август 2026</code> — месяц\n"
    "• <code>q3 2026</code> — квартал\n"
    "• <code>2026</code> — весь год\n"
    "• <code>01.01.2026-31.03.2026</code> — свои даты\n"
    "• <code>7 дней</code>, <code>вчера</code>, <code>прошлый месяц</code>"
)


def kb_after_period(target: str) -> InlineKeyboardMarkup:
    return kb_back(
        [
            [
                InlineKeyboardButton(
                    "\U0001f9fe Как посчитано", callback_data="check:calc"
                ),
                InlineKeyboardButton(
                    "\U0001f4c5 Другой период", callback_data=f"period:pick:{target}"
                ),
            ]
        ]
    )


async def show_dashboard(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    repo: "Repo",
    start: date,
    end: date,
) -> None:
    remember_period(context, start, end)
    span = max((end - start).days, 1)
    prev_start, prev_end = start - timedelta(days=span), start
    cur, before, breakdown, cats = await asyncio.gather(
        asyncio.to_thread(repo.totals, start, end - timedelta(days=1)),
        asyncio.to_thread(repo.totals, prev_start, prev_end - timedelta(days=1)),
        asyncio.to_thread(repo.breakdown, start, end),
        asyncio.to_thread(repo.categories, start, end),
    )
    label = period_label(start, end)
    lines, _ = calc_sheet(label, breakdown, cats)
    context.user_data["calc"] = lines
    await send(
        update,
        dashboard_text(
            label, period_label(prev_start, prev_end), span, cur, before, breakdown, cats
        ),
        kb_back(
            [
                [
                    InlineKeyboardButton(
                        "\U0001f916 Разбор от ИИ",
                        callback_data=f"report:range:{start.isoformat()}:{end.isoformat()}",
                    ),
                    InlineKeyboardButton(
                        "\U0001f9fe Как посчитано", callback_data="check:calc"
                    ),
                ],
                [
                    InlineKeyboardButton(
                        "\U0001f4c5 Другой период", callback_data="period:pick:dash"
                    ),
                    InlineKeyboardButton(
                        "\U0001f4e4 Выгрузить CSV",
                        callback_data=f"export:range:{start.isoformat()}:{end.isoformat()}",
                    ),
                ],
                [InlineKeyboardButton("\U0001f4ac Спросить бухгалтера", callback_data="ai:ask")],
            ]
        ),
    )


async def run_period_report(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    repo: "Repo",
    ai: LLMService,
    start: date,
    end: date,
) -> None:
    remember_period(context, start, end)
    label = period_label(start, end)
    await update.effective_chat.send_action(ChatAction.TYPING)
    breakdown, cats = await asyncio.gather(
        asyncio.to_thread(repo.breakdown, start, end),
        asyncio.to_thread(repo.categories, start, end),
    )
    if not breakdown:
        await send(update, f"За {label} операций нет.", kb_period("report"))
        return

    lines, _ = calc_sheet(label, breakdown, cats)
    context.user_data["calc"] = lines

    revenue = sum((r["revenue"] for r in breakdown), Decimal(0))
    expenses = sum((r["expenses"] for r in breakdown), Decimal(0))
    shares = sum((r["artist_share"] for r in breakdown), Decimal(0))
    profit = sum((r["label_profit"] for r in breakdown), Decimal(0))
    head = [
        f"\U0001f4ca <b>Отчёт · {label}</b>",
        f"Выручка: <b>{money(revenue, Config.DEFAULT_CURRENCY)}</b>",
        f"Расходы: <b>{money(expenses, Config.DEFAULT_CURRENCY)}</b>",
        f"Чистыми: <b>{money(revenue - expenses, Config.DEFAULT_CURRENCY)}</b> "
        f"<code>{money(revenue)} − {money(expenses)} = {money(revenue - expenses)}</code>",
        f"К выплате артистам: <b>{money(shares, Config.DEFAULT_CURRENCY)}</b>",
        f"Прибыль лейбла: <b>{money(profit, Config.DEFAULT_CURRENCY)}</b>",
        "",
        "<b>По артистам (с расчётом)</b>",
    ]
    for r in breakdown[:15]:
        head.append(
            f"• {r['artist']}: <code>({money(r['revenue'])} − {money(r['expenses'])}) "
            f"× {Decimal(str(r['rate'] or 0))} = {money(r['artist_share'])}</code> артисту, "
            f"лейблу {money(r['label_profit'])}"
        )
    await send(update, "\n".join(head))

    try:
        res = await ai.period_report(
            label,
            {"по_артистам": breakdown, "по_категориям": cats},
            lines,
            manual=chosen_model(context),
        )
    except RuntimeError as exc:
        await send(update, f"⚠️ {exc}", kb_main())
        return

    await asyncio.to_thread(
        repo.save_report,
        "period_report",
        res["text"],
        {"period": label, "breakdown": breakdown, "categories": cats},
        (start, end),
    )
    await send(
        update,
        f"\U0001f916 <b>Разбор ИИ · {label}</b>\n\n{res['text']}{route_footer(res)}",
        kb_after_period("report"),
    )


def kb_gen() -> InlineKeyboardMarkup:
    """Вкладка «Генерация»: две кнопки, обе без расхода токенов."""
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(
                    "\U0001f4c4 Генерация документа/отчёта", callback_data="gen:doc"
                )
            ],
            [
                InlineKeyboardButton(
                    "\U0001f5bc Генерация фотоотчёта", callback_data="gen:photo"
                )
            ],
            [InlineKeyboardButton("⬅️ Главное меню", callback_data="nav:main")],
        ]
    )


def kb_formats(domain: str, back: str) -> InlineKeyboardMarkup:
    rows = [
        [
            InlineKeyboardButton(
                f"{spec['icon']} {spec['title']}", callback_data=f"{domain}:fmt:{key}"
            )
        ]
        for key, spec in DOC_FORMATS.items()
    ]
    rows.append([InlineKeyboardButton("⬅️ Назад", callback_data=back)])
    return InlineKeyboardMarkup(rows)


def gen_menu_text() -> str:
    return (
        "\U0001f3a8 <b>Генерация</b>\n"
        "Файлы и картинки собирает код бота, а не модель — расход токенов нулевой.\n\n"
        "1. \U0001f4c4 <b>Документ/отчёт</b> — Excel с живыми формулами, Word или CSV.\n"
        "2. \U0001f5bc <b>Фотоотчёт</b> — картинка PNG с цифрами, графиком и формулами "
        "(без GPT Images, 0 токенов).\n\n"
        "<b>Токены ИИ</b>\n" + TOKENS.stats_text()
    )


def gen_format(context: ContextTypes.DEFAULT_TYPE) -> str:
    return str(context.user_data.get("gen_fmt") or "xlsx")


async def build_document(
    repo: "Repo", start: date, end: date, fmt: str
) -> tuple[bytes, str, str]:
    """Собирает файл по данным базы. Ни одного запроса к модели."""
    label = period_label(start, end)
    rows, cats, txs = await asyncio.gather(
        asyncio.to_thread(repo.breakdown, start, end),
        asyncio.to_thread(repo.categories, start, end),
        asyncio.to_thread(repo.period_transactions, start, end - timedelta(days=1)),
    )
    if not txs:
        raise ValueError("за период нет операций")

    calc_lines, _ = calc_sheet(label, rows, cats)
    if fmt == "csv":
        payload = await asyncio.to_thread(transactions_csv, txs)
        name = report_filename("finbot-operacii", start, end, "csv")
    elif fmt == "docx":
        blocks = report_blocks(label, rows, cats, calc_lines, len(txs))
        payload = await asyncio.to_thread(
            docx_report, f"Отчёт лейбла {Config.LABEL_NAME} · {label}", blocks
        )
        name = report_filename("finbot-otchet", start, end, "docx")
    elif fmt == "png":
        payload = await asyncio.to_thread(
            photo_report, label, rows, cats, len(txs), None
        )
        name = report_filename("finbot-fotootchet", start, end, "png")
    else:
        payload = await asyncio.to_thread(
            xlsx_report, label, rows, cats, txs, calc_lines
        )
        name = report_filename("finbot-otchet", start, end, "xlsx")

    totals = report_totals(rows)
    caption = (
        f"<b>{label}</b>\n"
        f"Выручка {money(totals['revenue'])} · расходы {money(totals['expenses'])} · "
        f"чистыми {money(totals['net'])}\n"
        f"Прибыль лейбла {money(totals['label_profit'])} · операций {len(txs)}\n"
        "<i>Считал код бота · токенов ИИ: 0</i>"
    )
    return payload, name, caption


async def deliver_document(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    repo: "Repo",
    start: date,
    end: date,
    fmt: str,
    target: str = "gen",
) -> None:
    remember_period(context, start, end)
    label = period_label(start, end)
    await update.effective_chat.send_action(ChatAction.TYPING)
    try:
        payload, name, caption = await build_document(repo, start, end, fmt)
    except ValueError:
        await send(
            update,
            f"За {label} операций нет — выгружать нечего. Выберите другой период.",
            kb_period(target),
        )
        return
    except RuntimeError as exc:
        await send(update, f"⚠️ {exc}", kb_gen())
        return
    except Exception as exc:  # noqa: BLE001 — любая ошибка сборки файла
        log.exception("document build failed")
        await send(update, f"⚠️ Не смог собрать файл: {exc}", kb_gen())
        return

    if fmt == "png":
        await update.effective_chat.send_photo(
            photo=InputFile(BytesIO(payload), filename=name),
            caption=f"\U0001f5bc <b>Фотоотчёт</b>\n{caption}",
            parse_mode=ParseMode.HTML,
        )
    else:
        spec = DOC_FORMATS.get(fmt, DOC_FORMATS["csv"])
        await update.effective_chat.send_document(
            document=InputFile(BytesIO(payload), filename=name),
            caption=f"{spec['icon']} <b>{spec['title']}</b>\n{caption}",
            parse_mode=ParseMode.HTML,
        )
    await show_main_menu(update)


async def export_period(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    repo: "Repo",
    start: date,
    end: date,
) -> None:
    """Экспорт документа: CSV, Excel или Word — формат выбирается кнопкой."""
    fmt = str(context.user_data.get("export_fmt") or "csv")
    await deliver_document(update, context, repo, start, end, fmt, target="export")


async def show_calc(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    repo: "Repo",
    start: date,
    end: date,
) -> None:
    """Экран «Как посчитано»: только арифметика кода, без ИИ."""
    remember_period(context, start, end)
    label = period_label(start, end)
    breakdown, cats = await asyncio.gather(
        asyncio.to_thread(repo.breakdown, start, end),
        asyncio.to_thread(repo.categories, start, end),
    )
    lines, _ = calc_sheet(label, breakdown, cats)
    context.user_data["calc"] = lines
    await send(update, calc_sheet_html(label, lines), kb_after_period("calc"))


async def show_compare(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    repo: "Repo",
    first: tuple[date, date],
    second: tuple[date, date],
) -> None:
    """Сравнение двух любых периодов — всё считает код."""
    (a_start, a_end), (b_start, b_end) = first, second
    a_totals, b_totals, a_rows, b_rows = await asyncio.gather(
        asyncio.to_thread(repo.totals, a_start, a_end - timedelta(days=1)),
        asyncio.to_thread(repo.totals, b_start, b_end - timedelta(days=1)),
        asyncio.to_thread(repo.breakdown, a_start, a_end),
        asyncio.to_thread(repo.breakdown, b_start, b_end),
    )
    a_label, b_label = period_label(a_start, a_end), period_label(b_start, b_end)
    a_share = sum((r["artist_share"] for r in a_rows), Decimal(0))
    b_share = sum((r["artist_share"] for r in b_rows), Decimal(0))

    def row(title: str, first_value: Decimal, second_value: Decimal) -> str:
        delta = second_value - first_value
        return (
            f"{title}: {money(first_value)} → <b>{money(second_value)}</b> "
            f"<code>{money(second_value)} − {money(first_value)} = {money(delta)}</code> "
            f"{pct_delta(second_value, first_value)}"
        )

    text = "\n".join(
        [
            f"⚖️ <b>Сравнение</b>\n{a_label} → {b_label}",
            "",
            row("Выручка", a_totals["revenue"], b_totals["revenue"]),
            row("Расходы", a_totals["expenses"], b_totals["expenses"]),
            row("Чистыми", a_totals["net"], b_totals["net"]),
            row("Артистам", a_share, b_share),
            "",
            f"Операций: {a_totals['count']} → {b_totals['count']}",
            "",
            "<i>Все разницы посчитаны кодом бота.</i>",
        ]
    )
    await send(
        update,
        text,
        kb_back([[InlineKeyboardButton("⚖️ Сравнить ещё", callback_data="cmp:menu")]]),
    )


async def apply_period(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    repo: "Repo",
    ai: LLMService,
    target: str,
    start: date,
    end: date,
) -> None:
    """Один выбор периода — много сценариев."""
    if target == "report":
        await run_period_report(update, context, repo, ai, start, end)
    elif target == "export":
        await export_period(update, context, repo, start, end)
    elif target == "gen":
        await deliver_document(
            update, context, repo, start, end, gen_format(context), target="gen"
        )
    elif target == "calc":
        await show_calc(update, context, repo, start, end)
    elif target == "cmpa":
        context.user_data["cmp_a"] = [start.isoformat(), end.isoformat()]
        await send(
            update,
            f"⚖️ Первый период: <b>{period_label(start, end)}</b>\n"
            "Шаг 2: выберите второй период для сравнения.",
            kb_period("cmpb"),
        )
    elif target == "cmpb":
        raw = context.user_data.get("cmp_a")
        if not raw:
            await send(update, "Начнём заново: выберите первый период.", kb_period("cmpa"))
            return
        first = (date.fromisoformat(raw[0]), date.fromisoformat(raw[1]))
        context.user_data.pop("cmp_a", None)
        await show_compare(update, context, repo, first, (start, end))
    else:
        await show_dashboard(update, context, repo, start, end)


async def show_audit(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    repo: "Repo",
    start: date,
    end: date,
) -> None:
    """Гигиена данных за период — без ИИ, чистая арифметика."""
    remember_period(context, start, end)
    label = period_label(start, end)
    rows, cats, txs = await asyncio.gather(
        asyncio.to_thread(repo.breakdown, start, end),
        asyncio.to_thread(repo.categories, start, end),
        asyncio.to_thread(repo.period_transactions, start, end - timedelta(days=1)),
    )
    no_artist = [r for r in rows if r["artist"] == NO_ARTIST]
    no_category = [c for c in cats if c["category"] == NO_CATEGORY]
    zero_rows = [t for t in txs if Decimal(str(t["amount"])) <= 0]
    seen: dict[tuple, int] = {}
    for t in txs:
        key = (str(t["occurred_on"])[:10], t["kind"], str(t["amount"]), t.get("artist"))
        seen[key] = seen.get(key, 0) + 1
    doubles = [key for key, count in seen.items() if count > 1]
    missing_rate = [
        r for r in rows if r["artist"] != NO_ARTIST and Decimal(str(r["rate"] or 0)) == 0
    ]

    lines = [
        f"\U0001fa7a <b>Проверка данных · {label}</b>",
        f"Всего операций: <b>{len(txs)}</b>",
        "",
        f"• Без артиста: <b>{sum(int(r['tx_count']) for r in no_artist)}</b>",
        f"• Без категории: <b>{sum(int(c['count']) for c in no_category)}</b>",
        f"• Нулевые или отрицательные суммы: <b>{len(zero_rows)}</b>",
        f"• Похожие на дубли: <b>{len(doubles)}</b>",
        f"• Артисты со ставкой 0%: <b>{len(missing_rate)}</b>",
    ]
    if doubles:
        lines += ["", "<b>Проверьте дубли:</b>"]
        for day, kind, amount, artist in doubles[:5]:
            sign = "доход" if kind == "income" else "расход"
            lines.append(f"• {day} · {sign} {amount} · {artist or 'без артиста'}")
    if missing_rate:
        lines += [
            "",
            "<b>Нет ставки — доля артиста считается как 0:</b>",
        ]
        lines += [f"• {r['artist']}" for r in missing_rate[:5]]
    lines += ["", "<i>Эти цифры посчитаны кодом, а не моделью.</i>"]
    await send(update, "\n".join(lines), kb_after_period("dash"))


async def quick_capture(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    repo: "Repo",
    ai: LLMService,
    text: str,
) -> None:
    """Любое сообщение в свободной форме → черновик операции с подтверждением."""
    artists = await asyncio.to_thread(repo.list_artists)
    await update.effective_chat.send_action(ChatAction.TYPING)
    try:
        parsed = await ai.quick_parse(
            text,
            [a["name"] for a in artists],
            date.today(),
            manual=chosen_model(context),
        )
    except Exception as exc:
        log.warning("quick_parse failed: %s", exc)
        await show_main_menu(update, f"⚠️ Не смог разобрать сообщение: {exc}")
        return

    if not parsed.get("understood") or not parsed.get("amount"):
        await show_main_menu(
            update,
            "🤔 Это не похоже на операцию. Напишите, например: "
            "<code>вчера зашло 120к роялти Spotify за MACAN</code> "
            "или <code>минус 15000 реклама в телеге</code>.",
        )
        return

    try:
        amount = parse_amount(str(parsed["amount"]))
    except ValueError as exc:
        await show_main_menu(update, f"⚠️ Сумма непонятна: {exc}")
        return

    try:
        occurred = parse_date(str(parsed.get("occurred_on") or ""))
    except ValueError:
        occurred = date.today()

    kind = "income" if str(parsed.get("kind")) == "income" else "expense"
    artist_name = parsed.get("artist") or None
    context.user_data["quick"] = {
        "kind": kind,
        "amount": str(amount),
        "category": parsed.get("category") or None,
        "description": (parsed.get("description") or text)[:400],
        "artist": artist_name,
        "occurred_on": occurred.isoformat(),
    }

    sign = "➕ Доход" if kind == "income" else "➖ Расход"
    note = parsed.get("note")
    confidence = parsed.get("confidence")
    tail = f"\n\n<i>{note}</i>" if note else ""
    if parsed.get("calc"):
        tail += f"\n\nСчёт: <code>{parsed['calc']}</code>"
    route = parsed.get("_route")
    if isinstance(route, dict):
        tail += route_footer(route)
    if isinstance(confidence, (int, float)) and confidence < 0.6:
        tail += "\n⚠️ Не всё однозначно — проверьте перед сохранением."

    await send(
        update,
        f"🧾 <b>Проверьте операцию</b>\n\n"
        f"{sign}: <b>{money(amount, Config.DEFAULT_CURRENCY)}</b>\n"
        f"Дата: {occurred:%d.%m.%Y}\n"
        f"Артист: {artist_name or '—'}\n"
        f"Категория: {parsed.get('category') or '—'}\n"
        f"Описание: {parsed.get('description') or '—'}{tail}",
        InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton("✅ Сохранить", callback_data="quick:save"),
                    InlineKeyboardButton("❌ Отмена", callback_data="nav:main"),
                ]
            ]
        ),
    )


# -----------------------------------------------------------------------------
# Текстовые сообщения (конечный автомат, без slash-команд)
# -----------------------------------------------------------------------------
async def on_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    repo: Repo = context.bot_data["repo"]
    ai: LLMService = context.bot_data["ai"]
    text = (update.effective_message.text or "").strip()

    access = await resolve_access(update, repo)
    if access is None:
        await handle_locked_message(update, repo, text)
        return

    state: str | None = context.user_data.get("state")

    if not state:
        # Без активного диалога: пробуем понять сообщение как операцию или вопрос.
        if len(text) < 4:
            await show_main_menu(update)
            return
        if not can_write(access["role"]):
            await update.effective_chat.send_action(ChatAction.TYPING)
            try:
                books = await collect_books(repo)
                answer = await ai.ask(text, books)
            except Exception as exc:
                await show_main_menu(update, f"⚠️ Не смог ответить: {exc}")
                return
            await send(update, answer, kb_main())
            return
        if any(ch.isdigit() for ch in text):
            await quick_capture(update, context, repo, ai, text)
            return
        await update.effective_chat.send_action(ChatAction.TYPING)
        try:
            books = await collect_books(repo)
            answer = await ai.ask(text, books)
        except Exception as exc:
            await show_main_menu(update, f"⚠️ Не смог ответить: {exc}")
            return
        await send(update, answer, kb_main())
        return

    try:
        if state == "artist:name":
            context.user_data["draft"] = {"name": text[:120]}
            context.user_data["state"] = "artist:rate"
            await send(
                update,
                f"Ставка артиста <b>{text[:120]}</b> в процентах (например <code>20</code>):",
                kb_back(),
            )

        elif state == "artist:rate":
            rate = parse_rate(text)
            draft = context.user_data.get("draft", {})
            artist = await asyncio.to_thread(
                repo.add_artist, draft["name"], rate, update.effective_user.username
            )
            await asyncio.to_thread(
                repo.audit, update.effective_user.id, "artist_add", artist
            )
            reset_state(context)
            await show_main_menu(
                update,
                f"✅ Артист <b>{artist['name']}</b> сохранён. "
                f"Ставка: {Decimal(str(artist['rate'])) * 100:.2f}%",
            )

        elif state == "tx:amount":
            amount = parse_amount(text)
            context.user_data["draft"]["amount"] = amount
            context.user_data["state"] = "tx:category"
            await send(
                update,
                f"Сумма: <b>{money(amount, Config.DEFAULT_CURRENCY)}</b>\n"
                "Теперь категория и описание одной строкой\n"
                "(например: <code>Стриминг | Spotify июль</code>):",
                kb_back(),
            )

        elif state == "tx:category":
            category, _, description = (x.strip() for x in text.partition("|"))
            context.user_data["draft"]["category"] = category[:80] or None
            context.user_data["draft"]["description"] = description[:400] or None
            context.user_data["state"] = "tx:date"
            await send(
                update, "Дата операции (ДД.ММ.ГГГГ) или «сегодня»:", kb_back()
            )

        elif state == "tx:date":
            draft = context.user_data["draft"]
            occurred = parse_date(text)
            await asyncio.to_thread(
                repo.add_transaction,
                draft.get("artist_id"),
                draft["kind"],
                draft["amount"],
                Config.DEFAULT_CURRENCY,
                draft.get("category"),
                draft.get("description"),
                occurred,
                "manual",
                None,
                update.effective_user.id,
            )
            label = "Доход" if draft["kind"] == "income" else "Расход"
            amount = draft["amount"]
            reset_state(context)
            await show_main_menu(
                update,
                f"✅ {label} {money(amount, Config.DEFAULT_CURRENCY)} "
                f"от {occurred:%d.%m.%Y} записан.",
            )

        elif state == "ai:ask":
            reset_state(context)
            await update.effective_chat.send_action(ChatAction.TYPING)
            start, end = current_period(context)
            books = await collect_books(repo, start, end)
            calc_lines = books_calc(books)
            context.user_data["calc"] = calc_lines
            res = await ai.ask(
                text,
                books,
                calc_lines,
                manual=chosen_model(context),
                hard=is_hard(text),
            )
            await send(
                update,
                f"💬 <b>Вопрос:</b> {text}\n"
                f"<i>Период: {period_label(start, end)}</i>\n\n"
                f"{res['text']}{route_footer(res)}",
                kb_after_period("dash"),
            )

        elif state == "period:custom":
            target = context.user_data.pop("period_target", "dash")
            reset_state(context)
            period = parse_period(text)
            if not period:
                await send(
                    update,
                    "Не понял период. " + CUSTOM_PERIOD_HINT,
                    kb_period(target),
                )
            else:
                await apply_period(
                    update, context, repo, ai, target, period[0], period[1]
                )

        elif state == "find:query":
            reset_state(context)
            found = await asyncio.to_thread(repo.search_transactions, text, 20)
            if not found:
                await send(
                    update,
                    f"🔍 По запросу «{text}» ничего не нашлось.",
                    kb_back([[InlineKeyboardButton(
                        "🔍 Искать ещё", callback_data="find:start"
                    )]]),
                )
            else:
                total_income = sum(
                    (r["amount"] for r in found if r["kind"] == "income"), Decimal(0)
                )
                total_expense = sum(
                    (r["amount"] for r in found if r["kind"] == "expense"), Decimal(0)
                )
                out = [f"🔍 <b>Найдено: {len(found)}</b> · «{text}»", ""]
                for r in found:
                    sign = "+" if r["kind"] == "income" else "−"
                    out.append(
                        f"• #{r['id']} {str(r['occurred_on'])[:10]} "
                        f"{sign}{money(r['amount'], Config.DEFAULT_CURRENCY)} · "
                        f"{r.get('artist') or 'без артиста'} · "
                        f"{r.get('category') or 'без категории'}"
                    )
                out += [
                    "",
                    f"Итого доходы: <b>{money(total_income)}</b>",
                    f"Итого расходы: <b>{money(total_expense)}</b>",
                    f"Разница: <code>{money(total_income)} − "
                    f"{money(total_expense)} = {money(total_income - total_expense)}</code>",
                ]
                await send(
                    update,
                    "\n".join(out),
                    kb_back([[InlineKeyboardButton(
                        "🔍 Искать ещё", callback_data="find:start"
                    )]]),
                )

        elif state == "access:grant":
            if not can_manage(access["role"]):
                raise ValueError("Нет прав на управление доступом.")
            fields = text.replace(",", " ").split()
            if not fields or not fields[0].lstrip("-").isdigit():
                raise ValueError(
                    "Нужен числовой Telegram ID, например 123456789 или 123456789 admin."
                )
            target_id = int(fields[0])
            new_role = fields[1].lower() if len(fields) > 1 else ROLE_VIEWER
            if new_role not in (ROLE_ADMIN, ROLE_VIEWER):
                raise ValueError("Роль может быть только admin или viewer.")
            if new_role == ROLE_ADMIN and access["role"] != ROLE_OWNER:
                raise ValueError("Админов назначает только владелец.")
            existing = await asyncio.to_thread(repo.get_user, target_id)
            if existing and existing["role"] == ROLE_OWNER:
                raise ValueError("Роль владельца менять нельзя.")
            await asyncio.to_thread(
                repo.grant, target_id, new_role, update.effective_user.id
            )
            await asyncio.to_thread(
                repo.audit,
                update.effective_user.id,
                "access_grant",
                {"target": target_id, "role": new_role},
            )
            reset_state(context)
            await show_main_menu(
                update,
                f"✅ Доступ выдан: <code>{target_id}</code> — {ROLE_TITLES[new_role]}.\n"
                "Попросите человека написать боту любое сообщение.",
            )

        else:
            reset_state(context)
            await show_main_menu(update)

    except ValueError as exc:
        await send(update, f"⚠️ {exc}\nПопробуйте ещё раз:", kb_back())


# -----------------------------------------------------------------------------
# Inline-кнопки
# -----------------------------------------------------------------------------
async def on_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()

    repo: Repo = context.bot_data["repo"]
    ai: LLMService = context.bot_data["ai"]

    access = await resolve_access(update, repo)
    if access is None:
        await query.edit_message_text(
            "⛔ Доступ закрыт. Попросите код-приглашение и пришлите его сообщением."
        )
        return
    role: str = access["role"]
    me: int = int(access["tg_user_id"])

    parts = (query.data or "").split(":")
    domain, action = parts[0], parts[1] if len(parts) > 1 else ""
    args = parts[2:]

    if (domain, action) in WRITE_GATED and not can_write(role):
        await send(update, "⛔ У вас доступ только для просмотра.", kb_main())
        return
    if domain == "access" and not can_manage(role):
        await send(
            update, "⛔ Управлять доступом могут владелец и админ.", kb_main()
        )
        return

    # ---------- навигация ----------
    if domain == "nav":
        reset_state(context)
        await show_main_menu(update)
        return

    # ---------- ИИ-бухгалтер: свободный вопрос ----------
    if domain == "ai":
        context.user_data["state"] = "ai:ask"
        await send(
            update,
            "💬 <b>Спросите ИИ-бухгалтера</b>\n\n"
            "Он видит все ваши операции, артистов, ставки и аномалии. Примеры:\n"
            "• <i>Сколько я должен артистам в этом месяце?</i>\n"
            "• <i>Почему расходы выросли?</i>\n"
            "• <i>Какой артист убыточен и почему?</i>\n"
            "• <i>Сколько отложить на налоги?</i>\n\n"
            "Напишите вопрос обычным сообщением.",
            kb_back(),
        )
        return

    # ---------- быстрый ввод из свободного текста ----------
    if domain == "quick":
        draft = context.user_data.get("quick")
        if action != "save" or not draft:
            reset_state(context)
            await show_main_menu(update, "Черновик устарел, напишите операцию заново.")
            return
        artist_id = None
        if draft.get("artist"):
            mapping = await asyncio.to_thread(
                repo.resolve_artist_ids, [draft["artist"]]
            )
            artist_id = mapping.get(str(draft["artist"]).strip().lower()) or next(
                iter(mapping.values()), None
            )
        amount = Decimal(draft["amount"])
        occurred = date.fromisoformat(draft["occurred_on"])
        await asyncio.to_thread(
            repo.add_transaction,
            artist_id,
            draft["kind"],
            amount,
            Config.DEFAULT_CURRENCY,
            draft.get("category"),
            draft.get("description"),
            occurred,
            "quick",
            None,
            me,
        )
        context.user_data.pop("quick", None)
        label = "Доход" if draft["kind"] == "income" else "Расход"
        await show_main_menu(
            update,
            f"✅ {label} {money(amount, Config.DEFAULT_CURRENCY)} "
            f"от {occurred:%d.%m.%Y} записан.",
        )
        return

    # ---------- сводка за любой период ----------
    if domain == "dash":
        if action == "period":
            await send(update, PERIOD_HINT, kb_period("dash"))
            return
        if action == "same":
            start, end = current_period(context)
        else:
            today = date.today()
            start, end = month_period(today.year, today.month)
        await show_dashboard(update, context, repo, start, end)
        return

    # ---------- экспорт в CSV ----------
    # ---------- генерация документов и фотоотчёта (без токенов) ----------
    if domain == "gen":
        if action == "menu":
            await send(update, gen_menu_text(), kb_gen())
            return
        if action == "doc":
            lines = [
                "\U0001f4c4 <b>Генерация документа</b>",
                "Выберите формат, потом период:",
                "",
            ]
            lines += [
                f"{spec['icon']} <b>{spec['title']}</b> — {spec['about']}"
                for spec in DOC_FORMATS.values()
            ]
            await send(update, "\n".join(lines), kb_formats("gen", "gen:menu"))
            return
        if action == "photo":
            context.user_data["gen_fmt"] = "png"
            await send(
                update,
                "\U0001f5bc <b>Фотоотчёт</b>\n"
                "Картинку рисует сам бот: крупные цифры, график статей, топ артистов "
                "и блок «Как посчитано». GPT Images не используется — расход токенов 0.\n\n"
                + PERIOD_HINT,
                kb_period("gen"),
            )
            return
        if action == "fmt":
            fmt = args[0] if args and args[0] in DOC_FORMATS else "xlsx"
            context.user_data["gen_fmt"] = fmt
            spec = DOC_FORMATS[fmt]
            await send(
                update,
                f"{spec['icon']} <b>{spec['title']}</b> — {spec['about']}\n\n"
                + PERIOD_HINT,
                kb_period("gen"),
            )
            return
        return

    if domain == "export":
        if action == "menu":
            lines = [
                "\U0001f4e4 <b>Экспорт документа</b>",
                "Сначала формат, потом период. Файлы собирает код бота — "
                "0 токенов ИИ.",
                "",
            ]
            lines += [
                f"{spec['icon']} <b>{spec['title']}</b> — {spec['about']}"
                for spec in DOC_FORMATS.values()
            ]
            await send(update, "\n".join(lines), kb_formats("export", "nav:main"))
            return
        if action == "fmt":
            fmt = args[0] if args and args[0] in DOC_FORMATS else "csv"
            context.user_data["export_fmt"] = fmt
            spec = DOC_FORMATS[fmt]
            await send(
                update,
                f"{spec['icon']} <b>{spec['title']}</b> — {spec['about']}\n\n"
                + PERIOD_HINT,
                kb_period("export"),
            )
            return
        if action == "month":
            start, end = month_period(int(args[0]), int(args[1]))
            await export_period(update, context, repo, start, end)
            return
        if action == "range":
            start, end = date.fromisoformat(args[0]), date.fromisoformat(args[1])
            await export_period(update, context, repo, start, end)
            return
        return

    # ---------- выбор периода: любой день, месяц, квартал, год ----------
    if domain == "period":
        target = args[0] if args else "dash"
        if action == "pick":
            await send(update, PERIOD_HINT, kb_period(target))
            return
        if action == "months":
            year = int(args[1]) if len(args) > 1 else date.today().year
            await send(
                update,
                f"\U0001f4c5 <b>{year} год</b>\nВыберите месяц или весь год:",
                kb_period_months(target, year),
            )
            return
        if action == "custom":
            context.user_data["state"] = "period:custom"
            context.user_data["period_target"] = target
            await send(update, CUSTOM_PERIOD_HINT, kb_back())
            return
        if action == "go" and len(args) >= 3:
            start, end = date.fromisoformat(args[1]), date.fromisoformat(args[2])
            await apply_period(update, context, repo, ai, target, start, end)
            return
        return

    # ---------- выбор модели ИИ ----------
    if domain == "model":
        if action == "auto":
            context.user_data["model_auto"] = True
            await send(
                update,
                "\U0001f916 Включён <b>автовыбор</b>: бот сам решает, кому считать, "
                "и поднимает модель, если цифры не сходятся.",
                kb_models(context),
            )
            return
        if action == "set" and args:
            key = normalize_model(args[0]) or base_model()
            spec = MODEL_CATALOG[key]
            context.user_data["model_auto"] = False
            context.user_data["model"] = key
            await send(
                update,
                f"✅ Зафиксирована модель <b>{spec['title']}</b> {spec['price']}\n"
                f"Лучше всего: {spec['best']}\n\n"
                "Автоэскалация в ручном режиме отключена.",
                kb_models(context),
            )
            return
        if action == "info" and args:
            key = normalize_model(args[0]) or base_model()
            spec = MODEL_CATALOG[key]
            await send(
                update,
                f"<b>{spec['title']}</b> {spec['price']} · <i>{TIER_TITLES[spec['tier']]}</i>\n\n"
                f"Лучше всего: {spec['best']}\n\n{spec['about']}",
                kb_models(context),
            )
            return
        mode = "автовыбор" if auto_mode(context) else "вручную"
        active = ACTIVE_MODEL.get("") or MODEL_CATALOG[base_model()]["title"]
        await send(
            update,
            f"\U0001f9e0 <b>Модели ИИ</b>\nРежим: <b>{mode}</b> · последняя считала: "
            f"<b>{active}</b>\nПровайдеры: {ai.provider_names()}\n\n"
            f"{AUTO_EXPLAIN}\n\n{models_help()}",
            kb_models(context),
        )
        return

    # ---------- «Как посчитано»: арифметика без ИИ ----------
    if domain == "check":
        start, end = current_period(context)
        await show_calc(update, context, repo, start, end)
        return

    # ---------- сравнение периодов ----------
    if domain == "cmp":
        context.user_data.pop("cmp_a", None)
        await send(
            update,
            "⚖️ <b>Сравнение периодов</b>\nШаг 1: выберите первый период.",
            kb_period("cmpa"),
        )
        return

    # ---------- поиск операций ----------
    if domain == "find":
        context.user_data["state"] = "find:query"
        await send(
            update,
            "\U0001f50d <b>Найти операцию</b>\n"
            "Напишите слово из описания, категорию или имя артиста — "
            "покажу все совпадения с суммами и датами.",
            kb_back(),
        )
        return

    # ---------- гигиена данных ----------
    if domain == "audit":
        start, end = current_period(context)
        await show_audit(update, context, repo, start, end)
        return

    # ---------- доступ и люди ----------
    if domain == "access":
        if action == "menu":
            users, invites = await asyncio.gather(
                asyncio.to_thread(repo.list_users),
                asyncio.to_thread(repo.list_invites),
            )
            active = [u for u in users if int(u["is_active"])]
            owners = [u for u in active if u["role"] == ROLE_OWNER]
            admins = [u for u in active if u["role"] == ROLE_ADMIN]
            viewers = [u for u in active if u["role"] == ROLE_VIEWER]
            await send(
                update,
                "🔑 <b>Доступ и люди</b>\n"
                f"Владельцы: <b>{len(owners)}</b>, админы: <b>{len(admins)}</b>, "
                f"наблюдатели: <b>{len(viewers)}</b>\n"
                f"Активных кодов-приглашений: <b>{len(invites)}</b>\n\n"
                "Кто что может:\n"
                "👑 <b>владелец</b> — всё, включая роли и удаление людей\n"
                "🛠 <b>админ</b> — вести финансы и звать наблюдателей\n"
                "👀 <b>наблюдатель</b> — только смотреть цифры и спрашивать ИИ",
                kb_access(role),
            )
            return

        if action == "users":
            users = await asyncio.to_thread(repo.list_users)
            if not users:
                await send(update, "Пока никого нет.", kb_access(role))
                return
            lines = []
            for u in users:
                mark = "" if int(u["is_active"]) else " 🚫 заблокирован"
                lines.append(
                    f"• <b>{user_title(u)}</b> — {ROLE_TITLES[u['role']]}{mark}\n"
                    f"  ID: <code>{u['tg_user_id']}</code>"
                )
            await send(
                update,
                "👥 <b>Кто имеет доступ</b>\n" + "\n".join(lines),
                kb_users(users),
            )
            return

        if action == "user":
            target = await asyncio.to_thread(repo.get_user, int(args[0]))
            if not target:
                await send(update, "Пользователь не найден.", kb_access(role))
                return
            status = "активен" if int(target["is_active"]) else "заблокирован"
            seen = target.get("last_seen_at") or "—"
            note = (
                ""
                if manageable(target, role, me)
                else "\n\n<i>Этого человека вы изменить не можете.</i>"
            )
            await send(
                update,
                f"👤 <b>{user_title(target)}</b>\n"
                f"Роль: {ROLE_TITLES[target['role']]}\n"
                f"Статус: {status}\n"
                f"ID: <code>{target['tg_user_id']}</code>\n"
                f"Последняя активность: {seen}{note}",
                kb_user(target, role, me),
            )
            return

        if action in ("role", "block", "unblock", "delete"):
            target = await asyncio.to_thread(repo.get_user, int(args[0]))
            if not target:
                await send(update, "Пользователь не найден.", kb_access(role))
                return
            if not manageable(target, role, me):
                await send(update, "⛔ Недостаточно прав для этого действия.", kb_access(role))
                return
            uid = int(target["tg_user_id"])

            if action == "role":
                new_role = args[1]
                if new_role not in (ROLE_ADMIN, ROLE_VIEWER) or role != ROLE_OWNER:
                    await send(update, "⛔ Роли меняет только владелец.", kb_access(role))
                    return
                await asyncio.to_thread(repo.set_role, uid, new_role)
                text_out = f"✅ Роль изменена: {ROLE_TITLES[new_role]}"
            elif action == "block":
                await asyncio.to_thread(repo.set_active, uid, False)
                text_out = "🚫 Доступ забран"
            elif action == "unblock":
                await asyncio.to_thread(repo.set_active, uid, True)
                text_out = "✅ Доступ возвращён"
            else:
                await asyncio.to_thread(repo.delete_user, uid)
                text_out = "🗑 Удалён из списка"

            await asyncio.to_thread(
                repo.audit, me, f"access_{action}", {"target": uid, "args": args}
            )
            users = await asyncio.to_thread(repo.list_users)
            await send(
                update,
                f"{text_out}: <b>{user_title(target)}</b>",
                kb_users(users),
            )
            return

        if action == "invite":
            invite_role = args[0] if args else ROLE_VIEWER
            if invite_role == ROLE_ADMIN and role != ROLE_OWNER:
                await send(update, "⛔ Код админа выдаёт только владелец.", kb_access(role))
                return
            inv = await asyncio.to_thread(
                repo.create_invite,
                invite_role,
                Config.INVITE_DEFAULT_USES,
                Config.INVITE_TTL_HOURS,
                me,
            )
            await asyncio.to_thread(repo.audit, me, "access_invite_create", inv)
            bot_username = getattr(context.bot, "username", None)
            link = (
                f"\nСсылка: <code>https://t.me/{bot_username}?start={inv['code']}</code>"
                if bot_username
                else ""
            )
            ttl = (
                f"действует {Config.INVITE_TTL_HOURS} ч"
                if Config.INVITE_TTL_HOURS > 0
                else "без срока"
            )
            await send(
                update,
                "🎫 <b>Код-приглашение готов</b>\n\n"
                f"<code>{inv['code']}</code>\n\n"
                f"Роль: {ROLE_TITLES[invite_role]}\n"
                f"Активаций: {inv['uses_left']}, {ttl}{link}\n\n"
                "Перешлите код человеку — он пришлёт его боту сообщением и сразу получит доступ.",
                kb_access(role),
            )
            return

        if action == "codes":
            invites = await asyncio.to_thread(repo.list_invites)
            if not invites:
                await send(update, "Активных кодов нет.", kb_access(role))
                return
            lines = [
                f"• <code>{i['code']}</code> — {ROLE_TITLES[i['role']]}, "
                f"активаций осталось: {i['uses_left']}"
                + (f", до {str(i['expires_at'])[:16].replace('T', ' ')}" if i["expires_at"] else "")
                for i in invites
            ]
            await send(
                update,
                "🔗 <b>Активные коды</b>\n" + "\n".join(lines),
                kb_codes(invites),
            )
            return

        if action == "kill":
            code = args[0] if args else ""
            await asyncio.to_thread(repo.revoke_invite, code)
            await asyncio.to_thread(repo.audit, me, "access_invite_revoke", {"code": code})
            invites = await asyncio.to_thread(repo.list_invites)
            await send(update, f"❌ Код <code>{code}</code> отозван.", kb_codes(invites))
            return

        if action == "grant":
            context.user_data["state"] = "access:grant"
            await send(
                update,
                "➕ <b>Выдать доступ по Telegram ID</b>\n\n"
                "Пришлите числовой ID, например <code>123456789</code> "
                "— по умолчанию роль наблюдателя.\n"
                "Чтобы сразу сделать админом: <code>123456789 admin</code>\n\n"
                "Свой ID человек увидит, если напишет этому боту любое сообщение.",
                kb_back(),
            )
            return
        return

    # ---------- артисты ----------
    if domain == "artist":
        if action == "add":
            context.user_data["state"] = "artist:name"
            await send(update, "👤 Введите имя (или название) артиста:", kb_back())
        elif action == "list":
            artists = await asyncio.to_thread(repo.list_artists)
            if not artists:
                await send(update, "Пока нет ни одного артиста.", kb_main())
                return
            lines = [
                f"• <b>{a['name']}</b> — {Decimal(str(a['rate'])) * 100:.2f}%"
                + (f" (@{a['tg_username']})" if a["tg_username"] else "")
                for a in artists
            ]
            await send(
                update, "🎤 <b>Артисты лейбла</b>\n" + "\n".join(lines), kb_main()
            )
        return

    # ---------- транзакции ----------
    if domain == "tx":
        if action == "new":
            kind = args[0]
            context.user_data["draft"] = {"kind": kind}
            context.user_data["state"] = None
            artists = await asyncio.to_thread(repo.list_artists)
            title = "➕ Доход" if kind == "income" else "➖ Расход"
            await send(
                update,
                f"{title}: выберите артиста",
                kb_artists(artists, f"tx:artist:{kind}", allow_none=True),
            )
        elif action == "undo":
            last = await asyncio.to_thread(repo.last_transaction_by, me)
            if not last:
                await send(update, "Вы ещё не добавляли операций.", kb_main())
                return
            label = "Доход" if last["kind"] == "income" else "Расход"
            await send(
                update,
                "↩️ <b>Отмена последней операции</b>\n\n"
                f"#{last['id']} • {label} "
                f"{money(last['amount'], Config.DEFAULT_CURRENCY)}\n"
                f"Дата: {str(last['occurred_on'])[:10]}\n"
                f"Артист: {last.get('artist') or '—'}\n"
                f"Категория: {last.get('category') or '—'}\n\n"
                "Удалить её?",
                InlineKeyboardMarkup(
                    [
                        [
                            InlineKeyboardButton(
                                "🗑 Удалить",
                                callback_data=f"tx:drop:{last['id']}",
                            ),
                            InlineKeyboardButton(
                                "❌ Оставить", callback_data="nav:main"
                            ),
                        ]
                    ]
                ),
            )
        elif action == "drop":
            tx_id = int(args[0])
            last = await asyncio.to_thread(repo.last_transaction_by, me)
            if not last or int(last["id"]) != tx_id:
                await show_main_menu(
                    update, "⚠️ Можно удалить только свою последнюю операцию."
                )
                return
            await asyncio.to_thread(repo.delete_transaction, tx_id)
            await asyncio.to_thread(repo.audit, me, "tx_delete", {"id": tx_id})
            await show_main_menu(update, f"🗑 Операция #{tx_id} удалена.")
        elif action == "artist":
            kind, artist_id = args[0], int(args[1])
            context.user_data["draft"] = {"kind": kind, "artist_id": artist_id or None}
            context.user_data["state"] = "tx:amount"
            await send(
                update,
                f"Введите сумму в {Config.DEFAULT_CURRENCY} (например <code>125000</code>):",
                kb_back(),
            )
        return

    # ---------- отчёты за любой период ----------
    if domain == "report":
        if action == "menu":
            await send(
                update,
                "\U0001f4ca <b>Отчёт с разбором ИИ</b>\n" + PERIOD_HINT,
                kb_period("report"),
            )
            return
        if action == "run" and len(args) >= 2:
            start, end = month_period(int(args[0]), int(args[1]))
            await run_period_report(update, context, repo, ai, start, end)
            return
        if action == "range" and len(args) >= 2:
            start, end = date.fromisoformat(args[0]), date.fromisoformat(args[1])
            await run_period_report(update, context, repo, ai, start, end)
            return
        return

    # ---------- Excel ----------
    if domain == "excel":
        context.user_data["state"] = None
        await send(
            update,
            "📥 <b>Импорт документа</b>\n"
            "Пришлите файл <code>.xlsx</code>, <code>.xls</code> или "
            "<code>.csv</code> как документ.\n\n"
            "Колонки в любом порядке и на любом языке: дата, артист, тип, "
            "сумма, категория, описание.\n"
            "Сначала таблицу разбирает код бота — бесплатно. ИИ подключается "
            "только к строкам, которые код не понял.",
            kb_back(),
        )
        return

    # ---------- аномалии ----------
    if domain == "anomaly":
        await update.effective_chat.send_action(ChatAction.TYPING)
        start, end = current_period(context)
        data, breakdown, cats = await asyncio.gather(
            asyncio.to_thread(repo.anomaly_candidates),
            asyncio.to_thread(repo.breakdown, start, end),
            asyncio.to_thread(repo.categories, start, end),
        )
        if not data["monthly"]:
            await send(update, "Недостаточно данных для анализа аномалий.", kb_main())
            return
        lines, _ = calc_sheet(period_label(start, end), breakdown, cats)
        context.user_data["calc"] = lines
        try:
            res = await ai.anomaly_report(data, lines, manual=chosen_model(context))
        except RuntimeError as exc:
            await send(update, f"⚠️ {exc}", kb_main())
            return
        await asyncio.to_thread(repo.save_report, "anomaly", res["text"], data, None)
        await send(
            update,
            f"\U0001f6a8 <b>Анализ аномалий</b>\n\n{res['text']}{route_footer(res)}",
            kb_after_period("dash"),
        )
        return

    # ---------- платежи ----------
    if domain == "pay":
        if action == "list":
            status = None if not args or args[0] == "all" else args[0]
            payments = await asyncio.to_thread(repo.list_payments, 20, status)
            if not payments:
                await send(update, "Платежей не найдено.", kb_payments([]))
                return
            icons = {"pending": "⏳", "paid": "✅", "canceled": "❌"}
            lines = ["💸 <b>Платежи артистам</b>"]
            for p in payments:
                lines.append(
                    f"{icons.get(p['status'], '•')} #{p['id']} <b>{p['artist']}</b> · "
                    f"{p['period_start']:%m.%Y} · "
                    f"{money(p['amount'], Config.DEFAULT_CURRENCY)} "
                    f"(ставка {p['rate'] * 100:.0f}%)"
                )
            total_pending = sum(
                (p["amount"] for p in payments if p["status"] == "pending"), Decimal(0)
            )
            lines.append(
                f"\nК выплате: <b>{money(total_pending, Config.DEFAULT_CURRENCY)}</b>"
            )
            await send(update, "\n".join(lines), kb_payments(payments))
        elif action == "recalc":
            today = date.today()
            created = await asyncio.to_thread(
                repo.upsert_payments_for_month, today.year, today.month
            )
            payments = await asyncio.to_thread(repo.list_payments, 20, "pending")
            await send(
                update,
                f"🔄 Пересчитано начислений: <b>{len(created)}</b> за {today:%m.%Y}.",
                kb_payments(payments),
            )
        elif action == "paid":
            payment_id = int(args[0])
            await asyncio.to_thread(repo.mark_payment_paid, payment_id)
            await asyncio.to_thread(
                repo.audit, update.effective_user.id, "payment_paid", {"id": payment_id}
            )
            payments = await asyncio.to_thread(repo.list_payments, 20, "pending")
            await send(
                update,
                f"✅ Платёж #{payment_id} отмечен как оплаченный.",
                kb_payments(payments),
            )
        return

    await show_main_menu(update)


# -----------------------------------------------------------------------------
# Excel-документы
# -----------------------------------------------------------------------------
def read_workbook(blob: bytes, max_rows: int | None = None) -> list[list[Any]]:
    max_rows = max_rows or Config.MAX_EXCEL_ROWS
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(blob), data_only=True, read_only=True)
    rows: list[list[Any]] = []
    try:
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                if row is None or all(cell in (None, "") for cell in row):
                    continue
                rows.append(
                    [
                        cell.strftime("%Y-%m-%d")
                        if isinstance(cell, (datetime, date))
                        else (float(cell) if isinstance(cell, Decimal) else cell)
                        for cell in row
                    ]
                )
                if len(rows) >= max_rows:
                    return rows
    finally:
        wb.close()
    return rows


# -----------------------------------------------------------------------------
# Бесплатный разбор таблиц без ИИ: что код понял сам — токены не тратим
# -----------------------------------------------------------------------------
CSV_DELIMS = (";", ",", "\t", "|")

HEADER_HINTS = {
    "date": ("дата", "date", "когда", "период", "месяц"),
    "artist": ("артист", "artist", "исполнитель", "псевдоним"),
    "kind": ("тип", "kind", "type", "вид", "операци", "дебет", "направлени"),
    "amount": ("сумма", "amount", "total", "итог", "выплат", "деньги", "value", "руб"),
    "category": ("категори", "category", "статья", "источник", "платформ"),
    "description": ("описани", "коммент", "note", "description", "назначени"),
}

INCOME_WORDS = (
    "доход",
    "income",
    "роялт",
    "royalt",
    "стрим",
    "stream",
    "концерт",
    "мерч",
    "синхро",
    "поступл",
    "приход",
    "выручк",
    "credit",
    "кредит",
)

EXPENSE_WORDS = (
    "расход",
    "expense",
    "реклам",
    "промо",
    "студи",
    "мастеринг",
    "микс",
    "дистриб",
    "аванс",
    "затрат",
    "списан",
    "оплата услуг",
    "debit",
    "дебет",
)

TOTAL_WORDS = ("итог", "total", "всего", "баланс", "subtotal")


def read_csv_blob(blob: bytes, max_rows: int | None = None) -> list[list[Any]]:
    """CSV с автоопределением кодировки и разделителя."""
    max_rows = max_rows or Config.MAX_EXCEL_ROWS
    text: str | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp1251", "latin-1"):
        try:
            text = blob.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("не смог определить кодировку файла")
    sample = text[:4000]
    delimiter = max(CSV_DELIMS, key=sample.count)
    rows: list[list[Any]] = []
    for row in csv.reader(StringIO(text), delimiter=delimiter):
        if not row or all(str(cell).strip() == "" for cell in row):
            continue
        rows.append([str(cell).strip() for cell in row])
        if len(rows) >= max_rows:
            break
    return rows


def _cell_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    if re.fullmatch(r"\d{4}-\d{1,2}-\d{1,2}", text):
        try:
            parts = [int(x) for x in text.split("-")]
            return date(parts[0], parts[1], parts[2])
        except ValueError:
            return None
    if re.fullmatch(r"\d{1,2}[./-]\d{1,2}[./-]\d{2,4}", text):
        try:
            return parse_date(text.replace("/", ".").replace("-", "."))
        except ValueError:
            return None
    if re.fullmatch(r"\d{1,2}[./-]\d{4}", text):
        month, year = re.split(r"[./-]", text)
        try:
            return date(int(year), int(month), 1)
        except ValueError:
            return None
    return None


def _cell_amount(value: Any) -> Decimal | None:
    """Сумма из любого вида записи: 15 000,50 / -15000.5 / (15000) / 15 000 ₽."""
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value)).quantize(CENTS, rounding=ROUND_HALF_UP)
        except (InvalidOperation, ValueError):
            return None
    text = str(value).strip()
    if not text or not re.search(r"\d", text):
        return None
    if _cell_date(text) is not None:
        return None
    digits = re.sub(r"[^\d\s.,\u00a0\u202f]", "", text)
    num = _norm_num(digits)
    if num is None:
        return None
    negative = text.startswith(("-", "\u2212")) or (
        text.startswith("(") and text.endswith(")")
    )
    return -num if negative else num


def sniff_columns(header: list[Any]) -> dict[str, int]:
    """Сопоставляет заголовки с полями — без запроса к модели."""
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header or []):
        text = str(cell or "").strip().lower()
        if not text:
            continue
        for field, hints in HEADER_HINTS.items():
            if field in mapping:
                continue
            if any(hint in text for hint in hints):
                mapping[field] = idx
                break
    return mapping


def local_parse(
    raw_rows: list[list[Any]],
) -> tuple[list[dict], list[str], list[list[Any]]]:
    """Разбор таблицы кодом. Возвращает (готовые строки, ошибки, непонятные)."""
    if not raw_rows:
        return [], [], []
    header = list(raw_rows[0])
    mapping = sniff_columns(header)
    if "amount" not in mapping or len(mapping) < 2:
        # Структура непонятна — отдаём всё модели.
        return [], [], [list(r) for r in raw_rows[1:]] or [list(r) for r in raw_rows]

    good: list[dict] = []
    errors: list[str] = []
    unparsed: list[list[Any]] = []

    for number, row in enumerate(raw_rows[1:], start=2):
        cells = list(row)

        def cell(field: str) -> Any:
            idx = mapping.get(field)
            if idx is None or idx >= len(cells):
                return None
            return cells[idx]

        blob_text = " ".join(
            str(c).lower() for c in cells if c not in (None, "")
        ).strip()
        if not blob_text:
            continue
        if any(word in blob_text for word in TOTAL_WORDS) and not _cell_date(
            cell("date")
        ):
            errors.append(f"строка {number}: похоже на итоговую — пропущена")
            continue

        amount = _cell_amount(cell("amount"))
        occurred = _cell_date(cell("date"))
        if occurred is None:
            for value in cells:
                occurred = _cell_date(value)
                if occurred is not None:
                    break
        if amount is None or occurred is None:
            unparsed.append(cells)
            continue

        kind: str | None = None
        kind_text = str(cell("kind") or "").lower()
        if any(word in kind_text for word in INCOME_WORDS):
            kind = "income"
        elif any(word in kind_text for word in EXPENSE_WORDS):
            kind = "expense"
        if kind is None and amount < 0:
            kind = "expense"
        if kind is None:
            hay = " ".join(
                str(cell(field) or "").lower() for field in ("category", "description")
            )
            if any(word in hay for word in EXPENSE_WORDS):
                kind = "expense"
            elif any(word in hay for word in INCOME_WORDS):
                kind = "income"
        if kind is None:
            unparsed.append(cells)
            continue

        amount = abs(amount)
        if amount <= 0:
            errors.append(f"строка {number}: нулевая сумма — пропущена")
            continue

        artist = str(cell("artist")).strip() if cell("artist") else None
        category = str(cell("category")).strip() if cell("category") else None
        description = str(cell("description")).strip() if cell("description") else None
        good.append(
            {
                "artist": artist or None,
                "kind": kind,
                "amount": str(amount),
                "currency": Config.DEFAULT_CURRENCY,
                "category": category or None,
                "description": description or None,
                "occurred_on": occurred.isoformat(),
            }
        )
    return good, errors, unparsed


async def on_document(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    repo: Repo = context.bot_data["repo"]
    ai: LLMService = context.bot_data["ai"]

    access = await resolve_access(update, repo)
    if access is None:
        return
    if not can_write(access["role"]):
        await send(
            update, "⛔ Импортировать документы могут владелец и админ.", kb_main()
        )
        return
    doc = update.effective_message.document
    file_name = (doc.file_name or "").lower()

    if not file_name.endswith((".xlsx", ".xlsm", ".xls", ".csv")):
        await send(
            update,
            "⚠️ Нужен документ Excel (<code>.xlsx</code> / <code>.xls</code>) "
            "или <code>.csv</code>.",
            kb_main(),
        )
        return
    if (doc.file_size or 0) > Config.MAX_EXCEL_BYTES:
        await send(update, "⚠️ Файл слишком большой.", kb_main())
        return

    await update.effective_chat.send_action(ChatAction.TYPING)
    tg_file = await doc.get_file()
    blob = bytes(await tg_file.download_as_bytearray())

    try:
        if file_name.endswith(".csv"):
            raw_rows = await asyncio.to_thread(read_csv_blob, blob)
        else:
            raw_rows = await asyncio.to_thread(read_workbook, blob)
    except Exception as exc:
        log.exception("import read failed")
        await send(update, f"⚠️ Не смог прочитать файл: {exc}", kb_main())
        return

    if not raw_rows:
        await send(update, "⚠️ Файл пустой.", kb_main())
        return

    local_rows, errors, leftovers = await asyncio.to_thread(local_parse, raw_rows)
    candidates: list[dict] = list(local_rows)
    summary = ""
    route: dict | None = None
    ai_rows = 0

    if leftovers:
        chunk = leftovers[: max(0, Config.AI_MAX_EXCEL_AI_ROWS)]
        if chunk:
            ai_rows = len(chunk)
            await send(
                update,
                f"\U0001f9ee Код бота разобрал бесплатно: <b>{len(local_rows)}</b> строк.\n"
                f"Отправляю модели только непонятные: <b>{ai_rows}</b>…",
            )
            try:
                parsed = await ai.parse_excel(
                    [list(raw_rows[0])] + chunk, manual=chosen_model(context)
                )
            except (RuntimeError, ValueError, json.JSONDecodeError) as exc:
                log.exception("excel parse failed")
                errors.append(f"ИИ не разобрал {ai_rows} строк: {exc}")
            else:
                candidates += list(parsed.get("rows", []) or [])
                errors += list(parsed.get("errors", []) or [])
                summary = str(parsed.get("summary") or "")
                if isinstance(parsed.get("_route"), dict):
                    route = parsed["_route"]
        if len(leftovers) > len(chunk):
            errors.append(
                f"не разобрано строк: {len(leftovers) - len(chunk)} "
                "(лимит AI_MAX_EXCEL_AI_ROWS для экономии токенов)"
            )
    else:
        await send(
            update,
            f"\U0001f9ee Разобрал файл без ИИ: <b>{len(local_rows)}</b> строк, 0 токенов.",
        )

    if not candidates:
        await send(
            update,
            "⚠️ Не нашёл ни одной операции. Нужны колонки с датой и суммой.",
            kb_main(),
        )
        return

    names = sorted(
        {str(r.get("artist")).strip() for r in candidates if r.get("artist")}
    )
    mapping = await asyncio.to_thread(repo.resolve_artist_ids, names)

    prepared: list[dict] = []
    for idx, r in enumerate(candidates):
        try:
            amount = Decimal(str(r["amount"])).quantize(CENTS, rounding=ROUND_HALF_UP)
            occurred = datetime.strptime(str(r["occurred_on"]), "%Y-%m-%d").date()
            kind = r["kind"]
            if kind not in ("income", "expense") or amount <= 0:
                raise ValueError("некорректный тип или сумма")
        except Exception as exc:
            errors.append(f"строка {idx + 1}: {exc}")
            continue

        artist_name = str(r.get("artist")).strip() if r.get("artist") else None
        artist_id = mapping.get(artist_name.lower()) if artist_name else None
        if artist_name and artist_id is None:
            errors.append(
                f"артист «{artist_name}» не найден в базе — операция без привязки"
            )

        prepared.append(
            {
                "artist_id": artist_id,
                "kind": kind,
                "amount": amount,
                "currency": (r.get("currency") or Config.DEFAULT_CURRENCY)[:3].upper(),
                "category": r.get("category") or None,
                "description": r.get("description") or None,
                "occurred_on": occurred,
                "external_key": (
                    f"xlsx:{doc.file_unique_id}:{idx}:{kind}:{to_cents(amount)}:{occurred}"
                ),
            }
        )

    inserted = await asyncio.to_thread(
        repo.bulk_insert_transactions, prepared, update.effective_user.id
    )
    await asyncio.to_thread(
        repo.save_report,
        "excel_import",
        summary,
        {
            "file": doc.file_name,
            "rows": len(prepared),
            "local": len(local_rows),
            "ai": ai_rows,
            "errors": errors,
        },
        None,
    )

    income = sum((r["amount"] for r in prepared if r["kind"] == "income"), Decimal(0))
    expense = sum((r["amount"] for r in prepared if r["kind"] == "expense"), Decimal(0))

    lines = [
        f"\U0001f4e5 <b>Импорт «{doc.file_name}» завершён</b>",
        f"Распознано строк: <b>{len(prepared)}</b>, записано новых: <b>{inserted}</b>",
        f"Без ИИ (бесплатно): <b>{len(local_rows)}</b> · через ИИ: <b>{ai_rows}</b>",
        f"Доходы: {money(income, Config.DEFAULT_CURRENCY)} · "
        f"Расходы: {money(expense, Config.DEFAULT_CURRENCY)}",
    ]
    if summary:
        lines += ["", f"<i>{summary}</i>"]
    if errors:
        lines += ["", "<b>⚠️ Проблемы в данных:</b>"] + [
            f"• {e}" for e in errors[:15]
        ]
        if len(errors) > 15:
            lines.append(f"…и ещё {len(errors) - 15}")
    if route:
        lines.append(route_footer(route))

    reset_state(context)
    await send(update, "\n".join(lines), kb_main())


# -----------------------------------------------------------------------------
# Глобальный обработчик ошибок
# -----------------------------------------------------------------------------
async def on_error(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    log.exception("Unhandled error", exc_info=context.error)
    if isinstance(update, Update) and update.effective_chat:
        try:
            await update.effective_chat.send_message(
                "⚠️ Внутренняя ошибка. Событие записано в лог, попробуйте ещё раз.",
                reply_markup=kb_main(),
            )
        except Exception:
            pass


# -----------------------------------------------------------------------------
# Точка входа (Bothost запускает именно этот файл)
# -----------------------------------------------------------------------------
def main() -> None:
    if not Config.BOT_TOKEN:
        raise SystemExit(
            "Не задана переменная окружения BOT_TOKEN. "
            "Добавьте её в разделе «Переменные окружения» панели Bothost."
        )
    if not active_providers():
        raise SystemExit(
            "Нужен хотя бы один ключ ИИ: XKIRO_API_KEY или TOOKEN_API_KEY. "
            "Добавьте его в разделе «Переменные окружения» панели Bothost."
        )

    db = Database()
    db.init_schema()
    repo = Repo(db)

    if Config.OWNER_IDS:
        repo.ensure_owners(Config.OWNER_IDS)
        log.info("Владельцы из OWNER_IDS: %s", sorted(Config.OWNER_IDS))
    elif repo.users_count() == 0:
        log.warning(
            "OWNER_IDS не задан: первый, кто напишет боту, станет владельцем"
        )
    ai = LLMService()

    builder = ApplicationBuilder().token(Config.BOT_TOKEN).concurrent_updates(True)
    if AIORateLimiter is not None:
        builder = builder.rate_limiter(AIORateLimiter())
    app = builder.build()
    app.bot_data["repo"] = repo
    app.bot_data["ai"] = ai

    app.add_handler(CallbackQueryHandler(on_callback))
    app.add_handler(MessageHandler(filters.Document.ALL, on_document))
    # Slash-команд нет: любой текст (включая /start) открывает главное меню
    app.add_handler(MessageHandler(filters.TEXT | filters.COMMAND, on_message))
    app.add_error_handler(on_error)

    log.info(
        "Bot started | build=%s | label=%s | mode=%s | base=%s | providers=%s | backend=%s",
        BUILD,
        Config.LABEL_NAME,
        Config.AI_MODE,
        MODEL_CATALOG[base_model()]["title"],
        ai.provider_names(),
        "postgres" if db.is_postgres else "sqlite",
    )
    try:
        app.run_polling(allowed_updates=Update.ALL_TYPES, drop_pending_updates=True)
    finally:
        db.close()


if __name__ == "__main__":
    main()
